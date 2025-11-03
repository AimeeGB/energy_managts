from decimal import Decimal
from rest_framework import views
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from django.http import JsonResponse

from .report_bitacora import generate_excel_report_bitacora
from .report_transformer_loss import generate_excel_report_transformer_loss
from .report_daily import generate_daily_report
# from .report_daily import DailyReportGenerator
from datetime import datetime
import openpyxl
from calendar import monthrange
import re

import os
from django.conf import settings
from io import BytesIO

from . import serializers
from . import models

class AreaLocalesAPIView(views.APIView):
    def get(self, request):
        # areas = models.AreaLocales.objects.filter(activo=True)
        areas = models.AreaLocales.objects.all()
        serializer = serializers.AreaLocalesSerializer(areas, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        serializer = serializers.AreaLocalesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        """Actualiza el estado activo/inactivo del área"""
        try:
            area = models.AreaLocales.objects.get(pk=pk)
            area.activo = not area.activo
            area.save()
            return Response({'activo': area.activo}, status=status.HTTP_200_OK)
        except models.AreaLocales.DoesNotExist:
            return Response({'error': 'Área no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        """Elimina permanentemente el área de la base de datos"""
        try:
            area = models.AreaLocales.objects.get(pk=pk)
            area.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except models.AreaLocales.DoesNotExist:
            return Response({'error': 'Área no encontrada'}, status=status.HTTP_404_NOT_FOUND)

#Codigo funcional 03/05/25
class EnergyConsumptionAPIView(views.APIView):
    def get(self, request):
        consumptions = models.EnergyConsumption.objects.all()
        serializer = serializers.EnergyConsumptionSerializer(consumptions, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        
        serializer = serializers.EnergyConsumptionSerializer(data=request.data)
        if serializer.is_valid():
            try:
                instance = serializer.create_or_update(serializer.validated_data)
                return Response(
                    serializers.EnergyConsumptionSerializer(instance).data,
                    status=status.HTTP_200_OK
                )
            except Exception as e:

                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request):
        # Obtener parámetros de identificación del request
        area = request.data.get('area')
        month = request.data.get('month')
        year = request.data.get('year')
        
        if not all([area, month, year]):
            return Response(
                {'error': 'Se requieren area, month y year para actualizar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            instance = models.EnergyConsumption.objects.get(area=area, month=month, year=year)
        except models.EnergyConsumption.DoesNotExist:
            return Response(
                {'error': 'Registro no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = serializers.EnergyConsumptionSerializer(instance, data=request.data, partial=False)
        if serializer.is_valid():
            try:
                instance = serializer.create_or_update(serializer.validated_data)
                return Response(
                    serializers.EnergyConsumptionSerializer(instance).data,
                    status=status.HTTP_200_OK
                )
            except Exception as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#Codigo de funcional 04/05/25       
class DownloadExcelAPIView(views.APIView):
    def get(self, request):
        try:
            month = int(request.query_params.get('month'))
            year = int(request.query_params.get('year'))
            
            # Get all instances for this month/year (each with their own plan_mes and perdidas)
            instances = models.EnergyConsumption.objects.filter(month=month, year=year)
            
            if not instances.exists():
                return Response(
                    {'error': 'No hay datos para el período especificado'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Use any instance (just need month/year)
            excel_file = generate_excel_report_bitacora(instances.first())
            
            # Nombre del mes en español
            month_names = {
                1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
                5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
                9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
            }

            month_name = month_names.get(month, "")
            response = HttpResponse(
                excel_file,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Tabla R 1A-Bitacora {month_name} {year}.xlsx"'
            return response
        except Exception as e:
            print(f"Error al generar Excel: {str(e)}")
            return Response(
                {'error': f'Error al generar el archivo Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

#Codigo de prueba 20/05/25
class GetEnergyDataAPIView(views.APIView):
    def get(self, request):
        area_id = request.query_params.get('area_id')
        month = int(request.query_params.get('month'))
        year = int(request.query_params.get('year'))
        
        if not area_id or not month or not year:
            return Response (
                {'error' : 'Se requiere el area, el mes y el año'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # area_id, month, year = int(area_id), int(month), int(year)
            area_id = int(area_id)
            month = int(month)
            year = int(year)
        except ValueError:
            return Response(
                {'error' : 'El area, el mes y el año deben ser numeros valido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            instance = models.EnergyConsumption.objects.get(
                area_id=area_id, 
                month=month, 
                year=year
            )
            serializer = serializers.EnergyConsumptionSerializer(instance)
            return Response(serializer.data)
        except models.EnergyConsumption.DoesNotExist:
            return Response(None)
        
class TransformerLossDataAPIView(views.APIView):
    def get(self, request):
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        
        if not all([month, year]):
            return Response(
                {'error': 'Both month and year parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        data = models.TransformerLossData.objects.filter(month=month, year=year)
        serializer = serializers.TransformerLossDataSerializer(data, many=True)
        return Response(serializer.data)
    
    def patch(self, request):
        # Primero validar que tenemos el ID para actualizar
        if 'id' not in request.data:
            return Response(
                {'error': 'Se requiere el ID del registro para actualizar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            instance = models.TransformerLossData.objects.get(id=request.data['id'])
        except models.TransformerLossData.DoesNotExist:
            return Response(
                {'error': 'No se encontró el registro para actualizar'},
                status=status.HTTP_404_NOT_FOUND
            )
        except KeyError:
            return Response(
                {'error': 'Se requiere el ID del registro para actualizar'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = serializers.TransformerLossDataSerializer(
            instance, 
            data=request.data, 
            partial=True
        )
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        serializer.save()
        
        return Response(serializer.data, status=status.HTTP_200_OK)


#Codigo de prueba 03/06/25
class ProcessBitacoraFileAPIView(views.APIView):    
    def post(self, request):
        month = request.data.get('month')
        year = request.data.get('year')
        
        if not all([month, year]):
            return Response(
                {'error': 'Month and year parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            month = int(month)
            year = int(year)
            
            # Find the file
            file_path = self.find_bitacora_file(month, year)
            if not file_path:
                return Response(
                    {'error': 'No se encontró el archivo TablaBitacora para el mes y año especificados'},
                    status=status.HTTP_404_NOT_FOUND
                )
                
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            results = []
            
            # Determine consumption cell based on days in month
            days_in_month = monthrange(year, month)[1]
            if days_in_month == 31:
                consumo_cell = 'Y11'
            elif month == 2 and days_in_month == 28:
                consumo_cell = 'Y14'
            elif month == 2 and days_in_month == 29:
                consumo_cell = 'Y13'
            else:
                consumo_cell = 'Y12'
                
            # Process each sheet in the workbook
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    # Get values
                    reactive_consumption = ws['M42'].value
                    monthly_consumption = ws[consumo_cell].value
                    
                    # Validación mejorada
                    if reactive_consumption in (None, '') or monthly_consumption in (None, ''):
                        results.append({
                            'area': sheet_name,
                            'status': 'error',
                            'error': f"Celdas requeridas (M42 o {consumo_cell}) están vacías o no son válidas"
                        })
                        continue
                        
                    try:
                        # Convertir a float, manejar posibles errores
                        reactive_consumption = float(reactive_consumption)
                        monthly_consumption = float(monthly_consumption)
                    except (ValueError, TypeError) as e:
                        results.append({
                            'area': sheet_name,
                            'status': 'error',
                            'error': f"Valores no numéricos en celdas (M42 o {consumo_cell})"
                        })
                        continue
                        
                    # Create or update record
                    data = {
                        'area': sheet_name,
                        'month': month,
                        'year': year,
                        'reactive_consumption': reactive_consumption,
                        'monthly_consumption': monthly_consumption,
                    }
                    
                    serializer = serializers.TransformerLossDataSerializer(data=data)
                    if serializer.is_valid():
                        instance = serializer.create_or_update(serializer.validated_data)
                        results.append({
                            'area': sheet_name,
                            'status': 'success',
                            'data': serializers.TransformerLossDataSerializer(instance).data
                        })
                    else:
                        results.append({
                            'area': sheet_name,
                            'status': 'error',
                            'error': serializer.errors
                        })
                        
                except Exception as e:
                    # logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                    results.append({
                        'area': sheet_name,
                        'status': 'error',
                        'error': str(e)
                    })
                    
            # Delete the file after processing
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Error al eliminar archivo: {str(e)}")
                
            return Response({
                'month': month,
                'year': year,
                'results': results
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            # logger.error(f"Error processing file: {str(e)}")
            return Response(
                {'error': f'Error al procesar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    def find_bitacora_file(self, month, year):
        """Busca el archivo TablaBitacora con patrones flexibles"""
        month_names = {
            1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
            5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
            9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
        }
        month_name = month_names.get(month, "")
        
        filename_patterns = [
            f"Tabla R 1A-Bitacora {month_name} {year}.xlsx",
            f"TablaBitacora {month_name} {year}.xlsx",
            f"Bitacora {month_name} {year}.xlsx"
        ]
        
        search_dirs = [
            os.path.join(settings.BASE_DIR, 'energeticos_app', 'generate_excel_report'),
            os.path.join(settings.MEDIA_ROOT, 'bitacora_files'),
            os.path.join(settings.BASE_DIR, 'temp_uploads')
        ]
        
        for directory in search_dirs:
            if not os.path.exists(directory):
                continue
                
            for pattern in filename_patterns:
                for file in os.listdir(directory):
                    if file.lower() == pattern.lower():
                        file_path = os.path.join(directory, file)
                        if os.path.exists(file_path):
                            return file_path
        return None


class DownloadExcelTransformerLossAPIView(views.APIView):
    def get(self, request):
        try:
            month = request.query_params.get('month')
            year = request.query_params.get('year')
            
            if not month or not year:
                return Response(
                    {'error': 'Month and year parameters are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            month = int(month)
            year = int(year)
            
            # Check if we have data for this month/year
            exists = models.TransformerLossData.objects.filter(
                month=month, year=year
            ).exists()
            
            if not exists:
                return Response(
                    {'error': 'No data available for the specified period'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Generate the report
            excel_file = generate_excel_report_transformer_loss(month, year)
            
            # Create response
            month_names = {
                1: "01 ENERO", 2: "02 FEBRERO", 3: "03 MARZO", 
                4: "04 ABRIL", 5: "05 MAYO", 6: "06 JUNIO",
                7: "07 JULIO", 8: "08 AGOSTO", 9: "09 SEPTIEMBRE",
                10: "10 OCTUBRE", 11: "11 NOVIEMBRE", 12: "12 DICIEMBRE"
            }
            month_name = month_names.get(month, "")
            # file_name = f"{month_name} {year}.xlsx"
            
            response = HttpResponse(
                excel_file,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Transformacion {month_name} {str(year)}.xlsx"'
            return response
            
        except ValueError:
            return Response(
                {'error': 'Invalid month or year parameter'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            print(f"Error generating Excel: {str(e)}")
            return Response(
                {'error': f'Error generating Excel file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# Añadir al views.py existente

# class DailyReportAPIView(views.APIView):
#     def get(self, request):
#         date_str = request.query_params.get('date')
#         if not date_str:
#             return Response(
#                 {'error': 'Se requiere parámetro date (YYYY-MM-DD)'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )
        
#         try:
#             selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
#             excel_file = generate_daily_report(selected_date)
            
#             response = HttpResponse(
#                 excel_file,
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = f'attachment; filename="Reporte_Diario_{selected_date}.xlsx"'
#             return response
            
#         except ValueError:
#             return Response(
#                 {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
#                 status=status.HTTP_400_BAD_REQUEST
#             )
#         except Exception as e:
#             return Response(
#                 {'error': f'Error al generar reporte: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )
        
        
# class AvailableDatesAPIView(views.APIView):
#     def get(self, request):
#         """Devuelve las fechas para las cuales hay datos disponibles"""
#         try:
#             # Obtener fechas que tienen datos en TablaBitacoraData
#             available_dates = models.TablaBitacoraData.objects.values_list(
#                 'year', 'month', 'day'
#             ).distinct()
            
#             dates = []
#             for year, month, day in available_dates:
#                 try:
#                     date_obj = datetime(year, month, day).date()
#                     dates.append(date_obj.isoformat())
#                 except ValueError:
#                     continue
            
#             return Response({'dates': sorted(dates)})
            
#         except Exception as e:
#             return Response(
#                 {'error': f'Error al obtener fechas disponibles: {str(e)}'},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR
#             )


import logging
logger = logging.getLogger(__name__)


# views.py - Modificar GenerateDailyReportAPIView
class GenerateDailyReportAPIView(views.APIView):
    def post(self, request):
        try:
            logger.info(f"Request data: {request.data}")
            date_str = request.data.get('date')
            if not date_str:
                return Response(
                    {'error': 'Fecha no proporcionada'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Verificar que el día no sea 0 para evitar división por cero
            if report_date.day == 0:
                return Response(
                    {'error': 'Fecha inválida: día no puede ser 0'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generar el reporte usando la función del módulo report_daily
            wb = generate_daily_report(report_date)
            
            # Guardar el archivo Excel
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Crear respuesta
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Reporte_Diario_{report_date.strftime("%Y-%m-%d")}.xlsx"'
            
            return response
            
        except ValueError as e:
            logger.error(f"Error de formato de fecha: {str(e)}")
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # Log del error completo para debugging
            import traceback
            error_trace = traceback.format_exc()
            logger.error(f"Error completo al generar reporte diario: {error_trace}")
            
            # Retornar el mensaje de error específico
            error_message = str(e)
            return Response(
                {'error': f'Error al generar el reporte: {error_message}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# En views.py, modificar la función que genera el reporte
# class GenerateDailyReportAPIView(views.APIView):
#     def post(self, request):
#         try:
#             date_str = request.data.get('date')
#             if not date_str:
#                 return Response({'error': 'Fecha requerida'}, status=status.HTTP_400_BAD_REQUEST)
            
#             report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
#             # Generar reporte con las modificaciones
#             wb = generate_daily_report(report_date)
            
#             # Guardar el archivo y devolverlo
#             response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#             response['Content-Disposition'] = f'attachment; filename=Reporte_Diario_{date_str}.xlsx'
            
#             wb.save(response)
#             return response
            
#         except Exception as e:
#             return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


# views.py
class AreaEstaticaAPIView(views.APIView):
    # def get(self, request):
    #     areas = models.AreaEstatica.objects.all().order_by('orden')
    #     serializer = serializers.AreaEstaticaSerializer(areas, many=True)
    #     return Response(serializer.data)
    
    # def post(self, request):
    #     serializer = serializers.AreaEstaticaSerializer(data=request.data)
    #     if serializer.is_valid():
    #         # Asignar el siguiente orden disponible
    #         if 'orden' not in request.data or not request.data['orden']:
    #             max_orden = models.AreaEstatica.objects.aggregate(max('orden'))['orden__max'] or 0
    #             serializer.validated_data['orden'] = max_orden + 1
                
    #         serializer.save()
    #         return Response(serializer.data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request):
        areas = models.AreaEstatica.objects.all().order_by('nombre')  # Ordenar por nombre en lugar de orden
        serializer = serializers.AreaEstaticaSerializer(areas, many=True)
        return Response(serializer.data)
    
    def post(self, request):
        serializer = serializers.AreaEstaticaSerializer(data=request.data)
        if serializer.is_valid():
            # Eliminar la lógica de orden automático
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def put(self, request, pk=None):
        try:
            area = models.AreaEstatica.objects.get(pk=pk)
        except models.AreaEstatica.DoesNotExist:
            return Response({'error': 'Área no encontrada'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = serializers.AreaEstaticaSerializer(area, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def delete(self, request, pk=None):
        try:
            area = models.AreaEstatica.objects.get(pk=pk)
            area.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except models.AreaEstatica.DoesNotExist:
            return Response({'error': 'Área no encontrada'}, status=status.HTTP_404_NOT_FOUND)

            