# #funcional - 22/09/25
# import os
# from datetime import datetime
# from decimal import Decimal
# import openpyxl
# from openpyxl.utils import get_column_letter
# from django.conf import settings
# from . import models
# from calendar import monthrange

# def generate_daily_report(report_date):
#     """
#     Función principal para generar el reporte diario
#     """
#     # Verificar áreas disponibles
#     areas = models.AreaLocales.objects.filter(
#         activo=True,
#         tablabitacoradata__month=report_date.month,
#         tablabitacoradata__year=report_date.year
#     ).distinct()
    
#     if not areas.exists():
#         raise Exception('No hay datos disponibles para el mes seleccionado')
    
#     # Cargar plantilla
#     template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloC.xlsx')
#     if not os.path.exists(template_path):
#         raise Exception('Plantilla de Excel no encontrada')
    
#     wb = openpyxl.load_workbook(template_path)
    
#     # Procesar hojas
#     modelo5_ws = wb['Modelo 5']
#     fill_modelo5_sheet(modelo5_ws, areas, report_date)
    
#     lecturas_ws = wb['Lecturas pico']
#     fill_peak_readings_sheet(lecturas_ws, areas, report_date)
    
#     return wb

# def fill_modelo5_sheet(worksheet, areas, report_date):
#     """
#     Llenar la hoja Modelo 5 con datos
#     """
#     # Procesar áreas estáticas (filas 8 a 25)
#     static_areas = models.AreaEstatica.objects.filter(activa=True)
#     static_row = 29
    
#     for area in static_areas:
#         worksheet[f'A{static_row}'] = static_row - 7  # Número consecutivo
#         worksheet[f'B{static_row}'] = area.provincia or ''  # Nueva columna provincia
#         worksheet[f'C{static_row}'] = area.municipio or ''  # Nueva columna municipio
#         worksheet[f'D{static_row}'] = area.nombre
#         worksheet[f'E{static_row}'] = area.plan_mes
        
#         # Ajustar fórmulas para las nuevas columnas
#         worksheet[f'F{static_row}'] = f'=J{static_row}*24'  # Plan acumulado (antes D)
#         worksheet[f'G{static_row}'] = f'=F{static_row}'     # Real acumulado (antes E)
#         worksheet[f'H{static_row}'] = f'=G{static_row}-F{static_row}'  # Real-Plan (antes F)
#         worksheet[f'I{static_row}'] = f'=IF(F{static_row}<>0, H{static_row}/F{static_row}, 0)'  # % (antes G)
#         worksheet[f'J{static_row}'] = f'=E{static_row}/31'  # Plan día (antes H)
#         worksheet[f'K{static_row}'] = f'=J{static_row}'     # Real día (antes I)
#         worksheet[f'L{static_row}'] = f'=K{static_row}-J{static_row}'  # Real-Plan (antes J)
#         worksheet[f'M{static_row}'] = f'=IF(J{static_row}<>0, L{static_row}/J{static_row}, 0)'  # % (antes K)
        
#         static_row += 1
    
#     # Determinar fila inicial para áreas dinámicas (continuando después de estáticas)
#     dynamic_start_row = static_row
    
#     # Procesar áreas dinámicas
#     dynamic_row = dynamic_start_row
#     dynamic_counter = static_row - 7
#     # dynamic_counter = dynamic_counter - 7
    
#     for area in areas:
#         try:
#             energy_consumption = models.EnergyConsumption.objects.get(
#                 area=area, 
#                 month=report_date.month, 
#                 year=report_date.year
#             )
            
#             bitacora_data = models.TablaBitacoraData.objects.get(
#                 area=area,
#                 month=report_date.month,
#                 year=report_date.year,
#                 day=report_date.day
#             )
            
#             # Calcular valores
#             perdidas = energy_consumption.perdidas or Decimal('0')
#             diarias = perdidas / report_date.day
            
#             consumo_mensual = bitacora_data.consumo_mensual or Decimal('0')
#             plan_mensual = bitacora_data.plan_mensual or Decimal('0')
#             consumo_acumulado = bitacora_data.consumo_acumulado or Decimal('0')
#             plan_acumulado = bitacora_data.plan_acumulado or Decimal('0')
            
#             # Escribir datos - ajustar columnas por las nuevas añadidas
#             worksheet[f'A{dynamic_row}'] = dynamic_counter  # Número consecutivo
#             worksheet[f'B{dynamic_row}'] = area.provincia  # Provincia (vacío para áreas dinámicas)
#             worksheet[f'C{dynamic_row}'] = area.municipio # Municipio (vacío para áreas dinámicas)
#             worksheet[f'D{dynamic_row}'] = area.nombre
#             worksheet[f'E{dynamic_row}'] = energy_consumption.plan_mes
#             worksheet[f'F{dynamic_row}'] = plan_acumulado
#             worksheet[f'G{dynamic_row}'] = consumo_acumulado + diarias
#             worksheet[f'J{dynamic_row}'] = plan_acumulado  # Plan día
#             worksheet[f'K{dynamic_row}'] = consumo_acumulado + diarias  # Real día
#             worksheet[f'N{dynamic_row}'] = perdidas  # Pérdidas
#             worksheet[f'O{dynamic_row}'] = diarias  # Diarias
            
#             # Establecer fórmulas
#             worksheet[f'H{dynamic_row}'] = f'=G{dynamic_row}-F{dynamic_row}'
#             worksheet[f'I{dynamic_row}'] = f'=IF(F{dynamic_row}<>0, H{dynamic_row}/F{dynamic_row}, 0)'
#             worksheet[f'L{dynamic_row}'] = f'=K{dynamic_row}-J{dynamic_row}'
#             worksheet[f'M{dynamic_row}'] = f'=IF(J{dynamic_row}<>0, L{dynamic_row}/J{dynamic_row}, 0)'
            
#             dynamic_row += 1
#             dynamic_counter += 1
            
#         except (models.EnergyConsumption.DoesNotExist, models.TablaBitacoraData.DoesNotExist):
#             continue
    
#     # Calcular totales
#     if dynamic_row > dynamic_start_row:  # Solo si hay datos
#         # total_row = dynamic_row
#         total_row = max(36, dynamic_row + 1)
#         first_static_row = 8
#         last_static_row = static_row - 1
#         first_dynamic_row = dynamic_start_row
        
#         worksheet[f'D{total_row}'] = 'TOTAL'
#         worksheet[f'E{total_row}'] = f'=SUM(E{first_static_row}:E{last_static_row}) + SUM(E{first_dynamic_row}:E{total_row-1})'
#         worksheet[f'F{total_row}'] = f'=SUM(F{first_static_row}:F{last_static_row}) + SUM(F{first_dynamic_row}:F{total_row-1})'
#         worksheet[f'G{total_row}'] = f'=SUM(G{first_static_row}:G{last_static_row}) + SUM(G{first_dynamic_row}:G{total_row-1})'
#         worksheet[f'H{total_row}'] = f'=G{total_row}-F{total_row}'
#         worksheet[f'I{total_row}'] = f'=IF(F{total_row}<>0, H{total_row}/F{total_row}, 0)'
#         worksheet[f'J{total_row}'] = f'=SUM(J{first_static_row}:J{last_static_row}) + SUM(J{first_dynamic_row}:J{total_row-1})'
#         worksheet[f'K{total_row}'] = f'=SUM(K{first_static_row}:K{last_static_row}) + SUM(K{first_dynamic_row}:K{total_row-1})'
#         worksheet[f'L{total_row}'] = f'=K{total_row}-J{total_row}'
#         worksheet[f'M{total_row}'] = f'=IF(J{total_row}<>0, L{total_row}/J{total_row}, 0)'

# # report_daily.py - Modificar fill_peak_readings_sheet
# def fill_peak_readings_sheet(worksheet, areas, report_date):
#     """
#     Llenar la hoja Lecturas Pico con datos - CORREGIDO
#     Cada área tiene su propia tabla independiente
#     """
#     # Configuración de columnas para cada tabla (CORREGIDO según el archivo original)
#     table_configs = [
#         {'start_col': 'B', 'area_name_col': 'B2', 'date_col': 'A'},
#         {'start_col': 'M', 'area_name_col': 'M2', 'date_col': 'L'},
#         {'start_col': 'W', 'area_name_col': 'W2', 'date_col': 'W'},
#         {'start_col': 'AH', 'area_name_col': 'AH2', 'date_col': 'AH'},
#         {'start_col': 'AS', 'area_name_col': 'AS2', 'date_col': 'AS'},
#         {'start_col': 'BD', 'area_name_col': 'BD2', 'date_col': 'BD'}
#     ]
    
#     # Obtener el último día del mes
#     _, last_day = monthrange(report_date.year, report_date.month)
    
#     # Procesar cada tabla
#     for i, config in enumerate(table_configs):
#         if i >= len(areas):
#             break
            
#         area = areas[i]
#         start_col = config['start_col']
#         date_col = config['date_col']
#         start_row = 5  # Fila donde empiezan los datos
        
#         # Escribir nombre del servicio desde la base de datos
#         worksheet[config['area_name_col']] = area.nombre
        
#         # Obtener todos los datos del mes para esta área
#         bitacora_data_list = models.TablaBitacoraData.objects.filter(
#             area=area,
#             month=report_date.month,
#             year=report_date.year
#         ).order_by('day')
        
#         # Llenar datos para cada día del mes
#         for day in range(1, last_day + 1):
#             try:
#                 # Buscar datos para este día
#                 bitacora_data = next((bd for bd in bitacora_data_list if bd.day == day), None)
                
#                 if bitacora_data:
#                     current_row = start_row + day - 1
                    
#                     # Escribir fecha en la columna correspondiente (CORREGIDO)
#                     worksheet[f"{date_col}{current_row}"] = datetime(report_date.year, report_date.month, day)
                    
#                     # Escribir lecturas (CORREGIDO según estructura del archivo original)
#                     # Columnas: B(1), C(2), D(3) para pico diurno; E(4), F(5), G(6) para pico nocturno; H(7), I(8), J(9) para consumo total
                    
#                     # PICO DIURNO
#                     worksheet[f"{get_next_col(start_col, 0)}{current_row}"] = bitacora_data.consumo_con_pico or 0  # Lectura 11.00 am
#                     worksheet[f"{get_next_col(start_col, 1)}{current_row}"] = bitacora_data.consumo_con_react or 0    # Lectura 1.00 pm
#                     worksheet[f"{get_next_col(start_col, 2)}{current_row}"] = f'={get_next_col(start_col, 1)}{current_row}-{get_next_col(start_col, 0)}{current_row}'  # CONSUMO diurno
                    
#                     # PICO NOCTURNO
#                     worksheet[f"{get_next_col(start_col, 3)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 5.00 pm
#                     worksheet[f"{get_next_col(start_col, 4)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 9.00 pm
#                     worksheet[f"{get_next_col(start_col, 5)}{current_row}"] = f'={get_next_col(start_col, 4)}{current_row}-{get_next_col(start_col, 3)}{current_row}'  # CONSUMO nocturno
                    
#                     # CONSUMO TOTAL
#                     worksheet[f"{get_next_col(start_col, 6)}{current_row}"] = bitacora_data.total_diario or 0    # Lectura INICIAL
#                     worksheet[f"{get_next_col(start_col, 7)}{current_row}"] = bitacora_data.total_diario or 0      # Lectura FINAL
#                     worksheet[f"{get_next_col(start_col, 8)}{current_row}"] = f'={get_next_col(start_col, 7)}{current_row}-{get_next_col(start_col, 6)}{current_row}'  # CONSUMO total
                    
#             except Exception as e:
#                 print(f"Error procesando día {day} para área {area.nombre}: {str(e)}")
#                 continue

# def get_date_column_for_table(table_index):
#     """
#     Obtener la columna de fecha para cada tabla según el índice
#     """
#     date_columns = ['A', 'L', 'W', 'AH', 'AS', 'BD']
#     return date_columns[table_index] if table_index < len(date_columns) else 'A'

# def get_next_col(current_col, offset):
#     """
#     Obtener la columna desplazada por offset
#     """
#     col_idx = openpyxl.utils.column_index_from_string(current_col) + offset
#     return openpyxl.utils.get_column_letter(col_idx)



###########################
import os
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings
from . import models
from calendar import monthrange

def generate_daily_report(report_date):
    """
    Función principal para generar el reporte diario
    """
    # Verificar áreas disponibles
    areas = models.AreaLocales.objects.filter(
        activo=True,
        tablabitacoradata__month=report_date.month,
        tablabitacoradata__year=report_date.year
    ).distinct()
    
    if not areas.exists():
        raise Exception('No hay datos disponibles para el mes seleccionado')
    
    # Cargar plantilla
    template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloC.xlsx')
    if not os.path.exists(template_path):
        raise Exception('Plantilla de Excel no encontrada')
    
    wb = openpyxl.load_workbook(template_path)
    
    # Procesar hojas
    modelo5_ws = wb['Modelo 5']
    fill_modelo5_sheet(modelo5_ws, areas, report_date)
    
    lecturas_ws = wb['Lecturas pico']
    fill_peak_readings_sheet(lecturas_ws, areas, report_date)
    
    return wb

def fill_modelo5_sheet(worksheet, areas, report_date):
    """
    Llenar la hoja Modelo 5 con datos
    """
    # Procesar áreas estáticas (filas 8 a 25)
    static_areas = models.AreaEstatica.objects.filter(activa=True)
    static_row = 29
    
    for area in static_areas:
        worksheet[f'A{static_row}'] = static_row - 7  # Número consecutivo
        worksheet[f'B{static_row}'] = area.provincia or ''  # Nueva columna provincia
        worksheet[f'C{static_row}'] = area.municipio or ''  # Nueva columna municipio
        worksheet[f'D{static_row}'] = area.nombre
        worksheet[f'E{static_row}'] = area.plan_mes
        
        # Ajustar fórmulas para las nuevas columnas
        worksheet[f'F{static_row}'] = f'=J{static_row}*24'  # Plan acumulado (antes D)
        worksheet[f'G{static_row}'] = f'=F{static_row}'     # Real acumulado (antes E)
        worksheet[f'H{static_row}'] = f'=G{static_row}-F{static_row}'  # Real-Plan (antes F)
        worksheet[f'I{static_row}'] = f'=IF(F{static_row}<>0, H{static_row}/F{static_row}, 0)'  # % (antes G)
        worksheet[f'J{static_row}'] = f'=E{static_row}/31'  # Plan día (antes H)
        worksheet[f'K{static_row}'] = f'=J{static_row}'     # Real día (antes I)
        worksheet[f'L{static_row}'] = f'=K{static_row}-J{static_row}'  # Real-Plan (antes J)
        worksheet[f'M{static_row}'] = f'=IF(J{static_row}<>0, L{static_row}/J{static_row}, 0)'  # % (antes K)
        
        static_row += 1
    
    # Determinar fila inicial para áreas dinámicas (continuando después de estáticas)
    dynamic_start_row = static_row

    # Filtrar áreas que tienen datos de EnergyConsumption para el mes y año del reporte
    areas_con_datos = []
    for area in areas:
        try:
            models.EnergyConsumption.objects.get(
                area=area, 
                month=report_date.month, 
                year=report_date.year
            )
            areas_con_datos.append(area)
        except models.EnergyConsumption.DoesNotExist:
            continue
    
    # Verificar si necesitamos insertar filas antes de agregar áreas dinámicas
    # total_estimated_rows = len(static_areas) + len(areas)
    total_estimated_rows = len(static_areas) + len(areas_con_datos)
    target_row_34 = 34  # Fila objetivo donde debe estar el total
    
    # Calcular cuántas filas necesitamos insertar
    # if dynamic_start_row + total_estimated_rows > target_row_34:
    #     rows_needed = (dynamic_start_row + total_estimated_rows) - target_row_34
    #     # Insertar filas después de la última fila estática
    #     worksheet.insert_rows(dynamic_start_row, amount=rows_needed)
        
    #     # Actualizar dynamic_start_row ya que las filas se han desplazado
    #     dynamic_start_row += rows_needed

    # Si la fila donde empezarían las dinámicas + las dinámicas supera la 34
    if dynamic_start_row + len(areas_con_datos) > target_row_34:
        rows_needed = (dynamic_start_row + len(areas_con_datos)) - target_row_34
        # Insertar filas después de la última fila estática
        worksheet.insert_rows(dynamic_start_row, amount=rows_needed)
        
        # Actualizar dynamic_start_row ya que las filas se han desplazado
        dynamic_start_row += rows_needed
    
    # Procesar áreas dinámicas
    dynamic_row = dynamic_start_row
    dynamic_counter = static_row - 7
    
    for area in areas:
        try:
            energy_consumption = models.EnergyConsumption.objects.get(
                area=area, 
                month=report_date.month, 
                year=report_date.year
            )
            
            bitacora_data = models.TablaBitacoraData.objects.get(
                area=area,
                month=report_date.month,
                year=report_date.year,
                day=report_date.day
            )
            
            # Calcular valores
            perdidas = energy_consumption.perdidas or Decimal('0')
            diarias = perdidas / report_date.day
            
            consumo_mensual = bitacora_data.consumo_mensual or Decimal('0')
            plan_mensual = bitacora_data.plan_mensual or Decimal('0')
            consumo_acumulado = bitacora_data.consumo_acumulado or Decimal('0')
            plan_acumulado = bitacora_data.plan_acumulado or Decimal('0')
            
            # Escribir datos - ajustar columnas por las nuevas añadidas
            worksheet[f'A{dynamic_row}'] = dynamic_counter  # Número consecutivo
            worksheet[f'B{dynamic_row}'] = area.provincia  # Provincia (vacío para áreas dinámicas)
            worksheet[f'C{dynamic_row}'] = area.municipio # Municipio (vacío para áreas dinámicas)
            worksheet[f'D{dynamic_row}'] = area.nombre
            worksheet[f'E{dynamic_row}'] = energy_consumption.plan_mes
            worksheet[f'F{dynamic_row}'] = plan_acumulado
            worksheet[f'G{dynamic_row}'] = consumo_acumulado + diarias
            worksheet[f'J{dynamic_row}'] = plan_acumulado  # Plan día
            worksheet[f'K{dynamic_row}'] = consumo_acumulado + diarias  # Real día
            worksheet[f'N{dynamic_row}'] = perdidas  # Pérdidas
            worksheet[f'O{dynamic_row}'] = diarias  # Diarias
            
            # Establecer fórmulas
            worksheet[f'H{dynamic_row}'] = f'=G{dynamic_row}-F{dynamic_row}'
            worksheet[f'I{dynamic_row}'] = f'=IF(F{dynamic_row}<>0, H{dynamic_row}/F{dynamic_row}, 0)'
            worksheet[f'L{dynamic_row}'] = f'=K{dynamic_row}-J{dynamic_row}'
            worksheet[f'M{dynamic_row}'] = f'=IF(J{dynamic_row}<>0, L{dynamic_row}/J{dynamic_row}, 0)'
            
            dynamic_row += 1
            dynamic_counter += 1
            
        except (models.EnergyConsumption.DoesNotExist, models.TablaBitacoraData.DoesNotExist):
            continue
    
    # Calcular totales - asegurarse de que esté en la fila 34
    if dynamic_row > dynamic_start_row:  # Solo si hay datos
        # total_row = 34  # Fija la fila total en la 34
        # total_row = dynamic_row
        total_row = max(36, dynamic_row)
        
        # Asegurarse de que la fila 34 esté limpia antes de escribir
        for col in range(1, 16):  # Columnas A a O
            col_letter = get_column_letter(col)
            worksheet[f'{col_letter}{total_row}'] = None
        
        first_static_row = 29
        last_static_row = static_row - 1
        first_dynamic_row = dynamic_start_row
        last_dynamic_row = dynamic_row - 1
        
        worksheet[f'D{total_row}'] = 'TOTAL'
        worksheet[f'E{total_row}'] = f'=SUM(E{first_static_row}:E{last_static_row}) + SUM(E{first_dynamic_row}:E{last_dynamic_row})'
        worksheet[f'F{total_row}'] = f'=SUM(F{first_static_row}:F{last_static_row}) + SUM(F{first_dynamic_row}:F{last_dynamic_row})'
        worksheet[f'G{total_row}'] = f'=SUM(G{first_static_row}:G{last_static_row}) + SUM(G{first_dynamic_row}:G{last_dynamic_row})'
        worksheet[f'H{total_row}'] = f'=G{total_row}-F{total_row}'
        worksheet[f'I{total_row}'] = f'=IF(F{total_row}<>0, H{total_row}/F{total_row}, 0)'
        worksheet[f'J{total_row}'] = f'=SUM(J{first_static_row}:J{last_static_row}) + SUM(J{first_dynamic_row}:J{last_dynamic_row})'
        worksheet[f'K{total_row}'] = f'=SUM(K{first_static_row}:K{last_static_row}) + SUM(K{first_dynamic_row}:K{last_dynamic_row})'
        worksheet[f'L{total_row}'] = f'=K{total_row}-J{total_row}'
        worksheet[f'M{total_row}'] = f'=IF(J{total_row}<>0, L{total_row}/J{total_row}, 0)'

# report_daily.py - Modificar fill_peak_readings_sheet
def fill_peak_readings_sheet(worksheet, areas, report_date):
    """
    Llenar la hoja Lecturas Pico con datos - CORREGIDO
    Cada área tiene su propia tabla independiente
    """
    # Configuración de columnas para cada tabla (CORREGIDO según el archivo original)
    table_configs = [
        {'start_col': 'B', 'area_name_col': 'B2', 'date_col': 'A'},
        {'start_col': 'M', 'area_name_col': 'M2', 'date_col': 'L'},
        {'start_col': 'W', 'area_name_col': 'W2', 'date_col': 'W'},
        {'start_col': 'AH', 'area_name_col': 'AH2', 'date_col': 'AH'},
        {'start_col': 'AS', 'area_name_col': 'AS2', 'date_col': 'AS'},
        {'start_col': 'BD', 'area_name_col': 'BD2', 'date_col': 'BD'}
    ]
    
    # Obtener el último día del mes
    _, last_day = monthrange(report_date.year, report_date.month)
    
    # Procesar cada tabla
    for i, config in enumerate(table_configs):
        if i >= len(areas):
            break
            
        area = areas[i]
        start_col = config['start_col']
        date_col = config['date_col']
        start_row = 5  # Fila donde empiezan los datos
        
        # Escribir nombre del servicio desde la base de datos
        worksheet[config['area_name_col']] = area.nombre
        
        # Obtener todos los datos del mes para esta área
        bitacora_data_list = models.TablaBitacoraData.objects.filter(
            area=area,
            month=report_date.month,
            year=report_date.year
        ).order_by('day')
        
        # Llenar datos para cada día del mes
        for day in range(1, last_day + 1):
            try:
                # Buscar datos para este día
                bitacora_data = next((bd for bd in bitacora_data_list if bd.day == day), None)
                
                if bitacora_data:
                    current_row = start_row + day - 1
                    
                    # Escribir fecha en la columna correspondiente (CORREGIDO)
                    worksheet[f"{date_col}{current_row}"] = datetime(report_date.year, report_date.month, day)
                    
                    # Escribir lecturas (CORREGIDO según estructura del archivo original)
                    # Columnas: B(1), C(2), D(3) para pico diurno; E(4), F(5), G(6) para pico nocturno; H(7), I(8), J(9) para consumo total
                    
                    # PICO DIURNO
                    worksheet[f"{get_next_col(start_col, 0)}{current_row}"] = bitacora_data.consumo_con_pico or 0  # Lectura 11.00 am
                    worksheet[f"{get_next_col(start_col, 1)}{current_row}"] = bitacora_data.consumo_con_react or 0    # Lectura 1.00 pm
                    worksheet[f"{get_next_col(start_col, 2)}{current_row}"] = f'={get_next_col(start_col, 1)}{current_row}-{get_next_col(start_col, 0)}{current_row}'  # CONSUMO diurno
                    
                    # PICO NOCTURNO
                    worksheet[f"{get_next_col(start_col, 3)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 5.00 pm
                    worksheet[f"{get_next_col(start_col, 4)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 9.00 pm
                    worksheet[f"{get_next_col(start_col, 5)}{current_row}"] = f'={get_next_col(start_col, 4)}{current_row}-{get_next_col(start_col, 3)}{current_row}'  # CONSUMO nocturno
                    
                    # CONSUMO TOTAL
                    worksheet[f"{get_next_col(start_col, 6)}{current_row}"] = bitacora_data.total_diario or 0    # Lectura INICIAL
                    worksheet[f"{get_next_col(start_col, 7)}{current_row}"] = bitacora_data.total_diario or 0      # Lectura FINAL
                    worksheet[f"{get_next_col(start_col, 8)}{current_row}"] = f'={get_next_col(start_col, 7)}{current_row}-{get_next_col(start_col, 6)}{current_row}'  # CONSUMO total
                    
            except Exception as e:
                print(f"Error procesando día {day} para área {area.nombre}: {str(e)}")
                continue


def get_date_column_for_table(table_index):
    """
    Obtener la columna de fecha para cada tabla según el índice
    """
    date_columns = ['A', 'L', 'W', 'AH', 'AS', 'BD']
    return date_columns[table_index] if table_index < len(date_columns) else 'A'


def get_next_col(current_col, offset):
    """
    Obtener la columna desplazada por offset
    """
    col_idx = openpyxl.utils.column_index_from_string(current_col) + offset
    return openpyxl.utils.get_column_letter(col_idx)





# import os
# from datetime import datetime
# from decimal import Decimal
# import openpyxl
# from openpyxl.utils import get_column_letter
# from django.conf import settings
# from . import models
# from calendar import monthrange

# def generate_daily_report(report_date):
#     """
#     Función principal para generar el reporte diario
#     """
#     # Verificar áreas disponibles
#     areas = models.AreaLocales.objects.filter(
#         activo=True,
#         tablabitacoradata__month=report_date.month,
#         tablabitacoradata__year=report_date.year
#     ).distinct()
    
#     if not areas.exists():
#         raise Exception('No hay datos disponibles para el mes seleccionado')
    
#     # Cargar plantilla
#     template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloC.xlsx')
#     if not os.path.exists(template_path):
#         raise Exception('Plantilla de Excel no encontrada')
    
#     wb = openpyxl.load_workbook(template_path)
    
#     # Procesar hojas
#     modelo5_ws = wb['Modelo 5']
#     fill_modelo5_sheet(modelo5_ws, areas, report_date)
    
#     lecturas_ws = wb['Lecturas pico']
#     fill_peak_readings_sheet(lecturas_ws, areas, report_date)
    
#     return wb

# def fill_modelo5_sheet(worksheet, areas, report_date):
#     """
#     Llenar la hoja Modelo 5 con datos
#     """
#     # Procesar áreas estáticas (filas 8 a 25)
#     static_areas = models.AreaEstatica.objects.filter(activa=True)
#     static_row = 29
    
#     for area in static_areas:
#         worksheet[f'A{static_row}'] = static_row - 7  # Número consecutivo
#         worksheet[f'B{static_row}'] = area.provincia or ''  # Nueva columna provincia
#         worksheet[f'C{static_row}'] = area.municipio or ''  # Nueva columna municipio
#         worksheet[f'D{static_row}'] = area.nombre
#         worksheet[f'E{static_row}'] = area.plan_mes
        
#         # Ajustar fórmulas para las nuevas columnas
#         worksheet[f'F{static_row}'] = f'=J{static_row}*24'  # Plan acumulado (antes D)
#         worksheet[f'G{static_row}'] = f'=F{static_row}'     # Real acumulado (antes E)
#         worksheet[f'H{static_row}'] = f'=G{static_row}-F{static_row}'  # Real-Plan (antes F)
#         worksheet[f'I{static_row}'] = f'=IF(F{static_row}<>0, H{static_row}/F{static_row}, 0)'  # % (antes G)
#         worksheet[f'J{static_row}'] = f'=E{static_row}/31'  # Plan día (antes H)
#         worksheet[f'K{static_row}'] = f'=J{static_row}'     # Real día (antes I)
#         worksheet[f'L{static_row}'] = f'=K{static_row}-J{static_row}'  # Real-Plan (antes J)
#         worksheet[f'M{static_row}'] = f'=IF(J{static_row}<>0, L{static_row}/J{static_row}, 0)'  # % (antes K)
        
#         static_row += 1
    
#     # Determinar fila inicial para áreas dinámicas (continuando después de estáticas)
#     dynamic_start_row = static_row

#     # Filtrar áreas que tienen datos de EnergyConsumption para el mes y año del reporte
#     areas_con_datos = []
#     for area in areas:
#         try:
#             models.EnergyConsumption.objects.get(
#                 area=area, 
#                 month=report_date.month, 
#                 year=report_date.year
#             )
#             areas_con_datos.append(area)
#         except models.EnergyConsumption.DoesNotExist:
#             continue
    
#     # Verificar si necesitamos insertar filas antes de agregar áreas dinámicas
#     target_row_34 = 34  # Fila objetivo donde debe estar el total
    
#     # Calcular cuántas filas necesitamos insertar
#     # CORRECCIÓN: Usar solo areas_con_datos para el cálculo
#     if dynamic_start_row + len(areas_con_datos) > target_row_34:
#         rows_needed = (dynamic_start_row + len(areas_con_datos)) - target_row_34
#         # Insertar filas después de la última fila estática
#         worksheet.insert_rows(dynamic_start_row, amount=rows_needed)
        
#         # Actualizar dynamic_start_row ya que las filas se han desplazado
#         dynamic_start_row += rows_needed
    
#     # CORRECCIÓN: Procesar solo áreas con datos, no todas las áreas
#     dynamic_row = dynamic_start_row
#     dynamic_counter = static_row - 7
    
#     for area in areas_con_datos:  # CAMBIO IMPORTANTE: usar areas_con_datos en lugar de areas
#         try:
#             energy_consumption = models.EnergyConsumption.objects.get(
#                 area=area, 
#                 month=report_date.month, 
#                 year=report_date.year
#             )
            
#             bitacora_data = models.TablaBitacoraData.objects.get(
#                 area=area,
#                 month=report_date.month,
#                 year=report_date.year,
#                 day=report_date.day
#             )
            
#             # Calcular valores
#             perdidas = energy_consumption.perdidas or Decimal('0')
#             diarias = perdidas / report_date.day
            
#             consumo_mensual = bitacora_data.consumo_mensual or Decimal('0')
#             plan_mensual = bitacora_data.plan_mensual or Decimal('0')
#             consumo_acumulado = bitacora_data.consumo_acumulado or Decimal('0')
#             plan_acumulado = bitacora_data.plan_acumulado or Decimal('0')
            
#             # Escribir datos
#             worksheet[f'A{dynamic_row}'] = dynamic_counter
#             worksheet[f'B{dynamic_row}'] = area.provincia
#             worksheet[f'C{dynamic_row}'] = area.municipio
#             worksheet[f'D{dynamic_row}'] = area.nombre
#             worksheet[f'E{dynamic_row}'] = energy_consumption.plan_mes
#             worksheet[f'F{dynamic_row}'] = plan_acumulado
#             worksheet[f'G{dynamic_row}'] = consumo_acumulado + diarias
#             worksheet[f'J{dynamic_row}'] = plan_acumulado  # Plan día
#             worksheet[f'K{dynamic_row}'] = consumo_acumulado + diarias  # Real día
#             worksheet[f'N{dynamic_row}'] = perdidas  # Pérdidas
#             worksheet[f'O{dynamic_row}'] = diarias  # Diarias
            
#             # Establecer fórmulas
#             worksheet[f'H{dynamic_row}'] = f'=G{dynamic_row}-F{dynamic_row}'
#             worksheet[f'I{dynamic_row}'] = f'=IF(F{dynamic_row}<>0, H{dynamic_row}/F{dynamic_row}, 0)'
#             worksheet[f'L{dynamic_row}'] = f'=K{dynamic_row}-J{dynamic_row}'
#             worksheet[f'M{dynamic_row}'] = f'=IF(J{dynamic_row}<>0, L{dynamic_row}/J{dynamic_row}, 0)'
            
#             dynamic_row += 1
#             dynamic_counter += 1
            
#         except (models.EnergyConsumption.DoesNotExist, models.TablaBitacoraData.DoesNotExist):
#             # Esto no debería pasar ya que filtramos areas_con_datos, pero por seguridad
#             continue
    
#     # Calcular totales - asegurarse de que esté en la fila 34
#     if dynamic_row > dynamic_start_row:  # Solo si hay datos
#         # total_row = 34  # Fija la fila total en la 34
#         total_row = dynamic_row
        
#         # Asegurarse de que la fila 34 esté limpia antes de escribir
#         for col in range(1, 16):  # Columnas A a O
#             col_letter = get_column_letter(col)
#             worksheet[f'{col_letter}{total_row}'] = None
        
#         first_static_row = 29
#         last_static_row = static_row - 1
#         first_dynamic_row = dynamic_start_row
#         last_dynamic_row = dynamic_row - 1
        
#         worksheet[f'D{total_row}'] = 'TOTAL'
#         worksheet[f'E{total_row}'] = f'=SUM(E{first_static_row}:E{last_static_row}) + SUM(E{first_dynamic_row}:E{last_dynamic_row})'
#         worksheet[f'F{total_row}'] = f'=SUM(F{first_static_row}:F{last_static_row}) + SUM(F{first_dynamic_row}:F{last_dynamic_row})'
#         worksheet[f'G{total_row}'] = f'=SUM(G{first_static_row}:G{last_static_row}) + SUM(G{first_dynamic_row}:G{last_dynamic_row})'
#         worksheet[f'H{total_row}'] = f'=G{total_row}-F{total_row}'
#         worksheet[f'I{total_row}'] = f'=IF(F{total_row}<>0, H{total_row}/F{total_row}, 0)'
#         worksheet[f'J{total_row}'] = f'=SUM(J{first_static_row}:J{last_static_row}) + SUM(J{first_dynamic_row}:J{last_dynamic_row})'
#         worksheet[f'K{total_row}'] = f'=SUM(K{first_static_row}:K{last_static_row}) + SUM(K{first_dynamic_row}:K{last_dynamic_row})'
#         worksheet[f'L{total_row}'] = f'=K{total_row}-J{total_row}'
#         worksheet[f'M{total_row}'] = f'=IF(J{total_row}<>0, L{total_row}/J{total_row}, 0)'

# # report_daily.py - Modificar fill_peak_readings_sheet
# def fill_peak_readings_sheet(worksheet, areas, report_date):
#     """
#     Llenar la hoja Lecturas Pico con datos - CORREGIDO
#     Cada área tiene su propia tabla independiente
#     """
#     # Configuración de columnas para cada tabla (CORREGIDO según el archivo original)
#     table_configs = [
#         {'start_col': 'B', 'area_name_col': 'B2', 'date_col': 'A'},
#         {'start_col': 'M', 'area_name_col': 'M2', 'date_col': 'L'},
#         {'start_col': 'W', 'area_name_col': 'W2', 'date_col': 'W'},
#         {'start_col': 'AH', 'area_name_col': 'AH2', 'date_col': 'AH'},
#         {'start_col': 'AS', 'area_name_col': 'AS2', 'date_col': 'AS'},
#         {'start_col': 'BD', 'area_name_col': 'BD2', 'date_col': 'BD'}
#     ]
    
#     # Obtener el último día del mes
#     _, last_day = monthrange(report_date.year, report_date.month)
    
#     # Procesar cada tabla
#     for i, config in enumerate(table_configs):
#         if i >= len(areas):
#             break
            
#         area = areas[i]
#         start_col = config['start_col']
#         date_col = config['date_col']
#         start_row = 5  # Fila donde empiezan los datos
        
#         # Escribir nombre del servicio desde la base de datos
#         worksheet[config['area_name_col']] = area.nombre
        
#         # Obtener todos los datos del mes para esta área
#         bitacora_data_list = models.TablaBitacoraData.objects.filter(
#             area=area,
#             month=report_date.month,
#             year=report_date.year
#         ).order_by('day')
        
#         # Llenar datos para cada día del mes
#         for day in range(1, last_day + 1):
#             try:
#                 # Buscar datos para este día
#                 bitacora_data = next((bd for bd in bitacora_data_list if bd.day == day), None)
                
#                 if bitacora_data:
#                     current_row = start_row + day - 1
                    
#                     # Escribir fecha en la columna correspondiente (CORREGIDO)
#                     worksheet[f"{date_col}{current_row}"] = datetime(report_date.year, report_date.month, day)
                    
#                     # Escribir lecturas (CORREGIDO según estructura del archivo original)
#                     # Columnas: B(1), C(2), D(3) para pico diurno; E(4), F(5), G(6) para pico nocturno; H(7), I(8), J(9) para consumo total
                    
#                     # PICO DIURNO
#                     worksheet[f"{get_next_col(start_col, 0)}{current_row}"] = bitacora_data.consumo_con_pico or 0  # Lectura 11.00 am
#                     worksheet[f"{get_next_col(start_col, 1)}{current_row}"] = bitacora_data.consumo_con_react or 0    # Lectura 1.00 pm
#                     worksheet[f"{get_next_col(start_col, 2)}{current_row}"] = f'={get_next_col(start_col, 1)}{current_row}-{get_next_col(start_col, 0)}{current_row}'  # CONSUMO diurno
                    
#                     # PICO NOCTURNO
#                     worksheet[f"{get_next_col(start_col, 3)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 5.00 pm
#                     worksheet[f"{get_next_col(start_col, 4)}{current_row}"] = bitacora_data.pico_diario or 0  # Lectura 9.00 pm
#                     worksheet[f"{get_next_col(start_col, 5)}{current_row}"] = f'={get_next_col(start_col, 4)}{current_row}-{get_next_col(start_col, 3)}{current_row}'  # CONSUMO nocturno
                    
#                     # CONSUMO TOTAL
#                     worksheet[f"{get_next_col(start_col, 6)}{current_row}"] = bitacora_data.total_diario or 0    # Lectura INICIAL
#                     worksheet[f"{get_next_col(start_col, 7)}{current_row}"] = bitacora_data.total_diario or 0      # Lectura FINAL
#                     worksheet[f"{get_next_col(start_col, 8)}{current_row}"] = f'={get_next_col(start_col, 7)}{current_row}-{get_next_col(start_col, 6)}{current_row}'  # CONSUMO total
                    
#             except Exception as e:
#                 print(f"Error procesando día {day} para área {area.nombre}: {str(e)}")
#                 continue

# def get_date_column_for_table(table_index):
#     """
#     Obtener la columna de fecha para cada tabla según el índice
#     """
#     date_columns = ['A', 'L', 'W', 'AH', 'AS', 'BD']
#     return date_columns[table_index] if table_index < len(date_columns) else 'A'

# def get_next_col(current_col, offset):
#     """
#     Obtener la columna desplazada por offset
#     """
#     col_idx = openpyxl.utils.column_index_from_string(current_col) + offset
#     return openpyxl.utils.get_column_letter(col_idx)

