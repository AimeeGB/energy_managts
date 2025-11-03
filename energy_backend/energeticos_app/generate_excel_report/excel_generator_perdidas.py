import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from django.conf import settings
from calendar import monthrange
import math
from .models import EnergyConsumption
from . import models

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center')
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_template_b():
    """Carga el archivo modeloB.xlsx desde los recursos de la energeticos_app"""
    import os
    from django.conf import settings
    
    # Ruta absoluta al archivo
    template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloB.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró el archivo modeloB en: {template_path}"
        )
    return openpyxl.load_workbook(template_path)

# def generate_excel_report_b(energy_consumption):
#     """Genera el reporte de pérdidas de transformación (modeloB.xlsx)"""
#     try:
#         # Cargar el modelo B
#         wb = load_template_b()
#         ws = wb['Pr']  # Asumiendo que la hoja se llama 'Pr'
        
#         # Configurar el mes y año en el título
#         month = energy_consumption.month
#         year = energy_consumption.year
        
#         month_names = {
#             1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
#             5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
#             9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
#         }
#         month_name = month_names.get(month, "")
        
#         # Actualizar título con mes y año
#         ws['B3'] = f"Cálculo de las pérdidas de transformación eléctrica. Mes {month_name} {year}"
        
#         # Actualizar título con mes y año (en negritas)
#         # cell_g2 = ws['B3']
#         # cell_g2.value = f"Cálculo de las pérdidas de transformación eléctrica. Mes {month_name} {year}"
#         # cell_g2.font = HEADER_FONT
#         # cell_g2.alignment = CENTER_ALIGN


#         # Configurar días según el mes
#         days = days_in_month(month, year)
#         ws['E7'] = days  # Días para Facultad de Ciencia Animal
#         ws['F7'] = days  # Días para Facultad de Ciencia Animal (segunda columna)
        
#         # Configurar horas (24 para todos)
#         ws['E8'] = 24  # Horas para Facultad de Ciencia Animal
#         ws['K8'] = 24  # Horas para Laboratorio Tec.Universidad
#         ws['Q8'] = 24  # Horas para CULTURA FISICA FAC. CAMAGUEY
#         ws['E23'] = 24  # Horas para Area Cultural
#         ws['K23'] = 24  # Horas para Instituto Superior Jose Marti
#         ws['E38'] = 24  # Horas para Cocina Universidad
#         ws['K38'] = 24  # Horas para Facultad Preparatoria
        
#         # Buscar el archivo TablaBitacora (puede estar en diferentes ubicaciones)
#         bitacora_path = find_bitacora_file(month, year)
#         if not bitacora_path:
#             raise FileNotFoundError("No se encontró el archivo TablaBitacora para este mes")
        
#         # Cargar el archivo TablaBitacora
#         bitacora_wb = openpyxl.load_workbook(bitacora_path, data_only=True)
        
#         # Mapeo de áreas a hojas en TablaBitacora y celdas en modeloB
#         areas_config = [
#             {
#                 'bitacora_sheet': 'Ciencia A',
#                 'consumo_mes_cell': 'E9',
#                 'consumo_react_cell': 'E10',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Laboratorios Tec.',
#                 'consumo_mes_cell': 'K9',
#                 'consumo_react_cell': 'K10',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Sede Fajardo',
#                 'consumo_mes_cell': 'Q9',
#                 'consumo_react_cell': 'Q10',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'A.Cultural',
#                 'consumo_mes_cell': 'E24',
#                 'consumo_react_cell': 'E25',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Sede Martí',
#                 'consumo_mes_cell': 'K24',
#                 'consumo_react_cell': 'K25',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Cocina C',
#                 'consumo_mes_cell': 'E39',
#                 'consumo_react_cell': 'E40',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Preparatoria',
#                 'consumo_mes_cell': 'K39',
#                 'consumo_react_cell': 'K40',
#                 'bitacora_consumo_cell': 'Y14',
#                 'bitacora_react_cell': 'M42'
#             }
#         ]
        
#         # Llenar datos de consumo para cada área
#         for area in areas_config:
#             try:
#                 bitacora_ws = bitacora_wb[area['bitacora_sheet']]
#                 consumo_mes = bitacora_ws[area['bitacora_consumo_cell']].value
#                 consumo_react = bitacora_ws[area['bitacora_react_cell']].value
                
#                 if consumo_mes is not None:
#                     ws[area['consumo_mes_cell']] = consumo_mes
#                 if consumo_react is not None:
#                     ws[area['consumo_react_cell']] = consumo_react
#             except Exception as e:
#                 print(f"Error al cargar datos para {area['bitacora_sheet']}: {str(e)}")
#                 continue
        
#         # Guardar el resultado en un buffer
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         return buffer.getvalue()
        
#     except Exception as e:
#         print(f"Error al generar Excel B: {str(e)}")
#         raise

# def generate_excel_report_b(energy_consumption):
#     """Genera el reporte de pérdidas de transformación (modeloB.xlsx)"""
#     try:
#         # Cargar el modelo B
#         wb = load_template_b()
#         ws = wb['Pr']  # Asumiendo que la hoja se llama 'Pr'

#         # Configurar el mes y año en el título
#         month = energy_consumption.month
#         year = energy_consumption.year
#         month_names = {
#             1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
#             5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
#             9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
#         }
#         month_name = month_names.get(month, "")

#         # Actualizar título con mes y año
#         ws['B3'] = f"Cálculo de las pérdidas de transformación eléctrica. Mes {month_name} {year}"

#         # Configurar días según el mes
#         days = days_in_month(month, year)
#         ws['E7'] = days  # Días para Facultad de Ciencia Animal
#         ws['F7'] = days  # Días para Facultad de Ciencia Animal (segunda columna)
#         ws['K7'] = days  # Días para Facultad de Laboratorio Tec.Universidad
#         ws['L7'] = days  # Días para Facultad de Laboratorio Tec.Universidad (segunda columna)
#         ws['Q7'] = days  # Días para Facultad de CULTURA FISICA FAC. CAMAGUEY
#         ws['R7'] = days  # Días para Facultad de CULTURA FISICA FAC. CAMAGUEY (segunda columna)
#         ws['E22'] = days  # Días para Area Cultural
#         ws['F22'] = days  # Días para Area Cultural (segunda columna)
#         ws['K22'] = days  # Días para Instituto Superior Jose Marti
#         ws['L22'] = days  # Días para Instituto Superior Jose Marti (segunda columna)
#         ws['E37'] = days  # Días para Cocina Universidad
#         ws['F37'] = days  # Días para Cocina Universidad (segunda columna)
#         ws['K37'] = days  # Días para Facultad Preparatoria
#         ws['L37'] = days  # Días para Facultad Preparatoria (segunda columna)

#         # Configurar horas (24 para todos)
#         ws['E8'] = 24  # Horas para Facultad de Ciencia Animal
#         ws['K8'] = 24  # Horas para Laboratorio Tec.Universidad
#         ws['Q8'] = 24  # Horas para CULTURA FISICA FAC. CAMAGUEY
#         ws['E23'] = 24  # Horas para Area Cultural
#         ws['K23'] = 24  # Horas para Instituto Superior Jose Marti
#         ws['E38'] = 24  # Horas para Cocina Universidad
#         ws['K38'] = 24  # Horas para Facultad Preparatoria

#         # Buscar el archivo TablaBitacora
#         bitacora_path = find_bitacora_file(month, year)
#         if not bitacora_path:
#             raise FileNotFoundError("No se encontró el archivo TablaBitacora para este mes")

#         # Cargar el archivo TablaBitacora
#         bitacora_wb = openpyxl.load_workbook(bitacora_path, data_only=True)

#         # Determinar la fila de consumo según el número de días del mes
#         if days == 31:
#             consumo_cell_row = 'Y11'  # Por defecto para meses de 31 días
#         # elif days == 30 and month == 6:
#         #     consumo_cell_row = 'Y12'
#         elif days == 28 and month == 2:  # Febrero
#             consumo_cell_row = 'Y14'
#         elif days == 29 and month == 2:  # Febrero
#             consumo_cell_row = 'Y13'
#         else:
#             consumo_cell_row = 'Y12'

#         # Mapeo de áreas a hojas en TablaBitacora y celdas en modeloB
#         areas_config = [
#             {
#                 'bitacora_sheet': 'Ciencia A',
#                 'consumo_mes_cell': 'E9',
#                 'consumo_react_cell': 'E10',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Laboratorios Tec.',
#                 'consumo_mes_cell': 'K9',
#                 'consumo_react_cell': 'K10',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Sede Fajardo',
#                 'consumo_mes_cell': 'Q9',
#                 'consumo_react_cell': 'Q10',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'A.Cultural',
#                 'consumo_mes_cell': 'E24',
#                 'consumo_react_cell': 'E25',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Sede Martí',
#                 'consumo_mes_cell': 'K24',
#                 'consumo_react_cell': 'K25',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Cocina C',
#                 'consumo_mes_cell': 'E39',
#                 'consumo_react_cell': 'E40',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             },
#             {
#                 'bitacora_sheet': 'Preparatoria',
#                 'consumo_mes_cell': 'K39',
#                 'consumo_react_cell': 'K40',
#                 'bitacora_consumo_cell': consumo_cell_row,
#                 'bitacora_react_cell': 'M42'
#             }
#         ]

#         # Llenar datos de consumo para cada área
#         for area in areas_config:
#             try:
#                 bitacora_ws = bitacora_wb[area['bitacora_sheet']]
#                 consumo_mes = bitacora_ws[area['bitacora_consumo_cell']].value
#                 consumo_react = bitacora_ws[area['bitacora_react_cell']].value

#                 if consumo_mes is not None:
#                     ws[area['consumo_mes_cell']] = consumo_mes
#                 if consumo_react is not None:
#                     ws[area['consumo_react_cell']] = consumo_react
#             except Exception as e:
#                 print(f"Error al cargar datos para {area['bitacora_sheet']}: {str(e)}")
#                 continue

#         # Guardar el resultado en un buffer
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         return buffer.getvalue()

#     except Exception as e:
#         print(f"Error al generar Excel B: {str(e)}")
#         raise

def generate_excel_report_transformer_loss(month, year):
    """Genera el reporte de pérdidas de transformación (modeloB.xlsx)"""
    try:
        # Cargar el modelo B
        wb = load_template_b()
        ws = wb['Pr']
        
        # Configurar el mes y año en el título
        month_names = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 
            4: "ABRIL", 5: "MAYO", 6: "JUNIO",
            7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
            10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
        month_name = month_names.get(month, "")
        ws['B3'] = f"Cálculo de las pérdidas de transformación eléctrica. Mes {month_name} {year}"
        
        # Configurar días según el mes
        days = days_in_month(month, year)
        ws['E7'] = days  # Días para Facultad de Ciencia Animal
        ws['F7'] = days  # Días para Facultad de Ciencia Animal (segunda columna)
        ws['K7'] = days  # Días para Laboratorio Tec.Universidad
        ws['L7'] = days  # Días para Laboratorio Tec.Universidad (segunda columna)
        ws['Q7'] = days  # Días para CULTURA FISICA FAC. CAMAGUEY
        ws['R7'] = days  # Días para CULTURA FISICA FAC. CAMAGUEY (segunda columna)
        ws['E22'] = days  # Días para Area Cultural
        ws['F22'] = days  # Días para Area Cultural (segunda columna)
        ws['K22'] = days  # Días para Instituto Superior Jose Marti
        ws['L22'] = days  # Días para Instituto Superior Jose Marti (segunda columna)
        ws['E37'] = days  # Días para Cocina Universidad
        ws['F37'] = days  # Días para Cocina Universidad (segunda columna)
        ws['K37'] = days  # Días para Facultad Preparatoria
        ws['L37'] = days  # Días para Facultad Preparatoria (segunda columna)
        
        # Configurar horas (24 para todos)
        ws['E8'] = 24  # Horas para Facultad de Ciencia Animal
        ws['K8'] = 24  # Horas para Laboratorio Tec.Universidad
        ws['Q8'] = 24  # Horas para CULTURA FISICA FAC. CAMAGUEY
        ws['E23'] = 24  # Horas para Area Cultural
        ws['K23'] = 24  # Horas para Instituto Superior Jose Marti
        ws['E38'] = 24  # Horas para Cocina Universidad
        ws['K38'] = 24  # Horas para Facultad Preparatoria
        
        # Mapeo de áreas a celdas en modeloB
        areas_config = [
            {
                'area': 'Ciencia A',
                'consumo_mes_cell': 'E9',
                'consumo_react_cell': 'E10',
                'kvan_cell': 'C12',
                'kvan_value': 167*3,  # 167 kVA * 3 (trifásico)
                'pfe_cell': 'C7',
                'pfe_value': 0.482*3,  # Pfe * 3 (trifásico)
                'pcu_cell': 'C8',
                'pcu_value': 1.893*3,  # Pcu * 3 (trifásico)
                'pr_cell': 'C17'
            },
            {
                'area': 'Laboratorios Tec.',
                'consumo_mes_cell': 'K9',
                'consumo_react_cell': 'K10',
                'kvan_cell': 'I12',
                'kvan_value': 25*3,  # 25 kVA * 3 (trifásico)
                'pfe_cell': 'I7',
                'pfe_value': 0.115*3,  # Pfe * 3 (trifásico)
                'pcu_cell': 'I8',
                'pcu_value': 0.389,  # Pcu (no multiplicar por 3)
                'pr_cell': 'I17'
            },
            {
                'area': 'Sede Fajardo',
                'consumo_mes_cell': 'Q9',
                'consumo_react_cell': 'Q10',
                'kvan_cell': 'O12',
                'kvan_value': 150,  # Valor fijo
                'pfe_cell': 'O7',
                'pfe_value': 0.83 + 0.66,  # Suma de dos transformadores
                'pcu_cell': 'O8',
                'pcu_value': 3.587 + 2.802,  # Suma de dos transformadores
                'pr_cell': 'O17'
            },
            {
                'area': 'A.Cultural',
                'consumo_mes_cell': 'E24',
                'consumo_react_cell': 'E25',
                'kvan_cell': 'C27',
                'kvan_value': 37.5*3,  # 37.5 kVA * 3 (trifásico)
                'pfe_cell': 'C22',
                'pfe_value': 0.162*3,  # Pfe * 3 (trifásico)
                'pcu_cell': 'C23',
                'pcu_value': 0.487*3,  # Pcu * 3 (trifásico)
                'pr_cell': 'C32'
            },
            {
                'area': 'Sede Martí',
                'consumo_mes_cell': 'K24',
                'consumo_react_cell': 'K25',
                'kvan_cell': 'I27',
                'kvan_value': 150,  # Valor fijo
                'pfe_cell': 'I22',
                'pfe_value': 0.199 + 0.332,  # Suma de dos transformadores
                'pcu_cell': 'I23',
                'pcu_value': 0.626 + 1.893,  # Suma de dos transformadores
                'pr_cell': 'I32'
            },
            {
                'area': 'Cocina C',
                'consumo_mes_cell': 'E39',
                'consumo_react_cell': 'E40',
                'kvan_cell': 'C42',
                'kvan_value': 50*3,  # 50 kVA * 3 (trifásico)
                'pfe_cell': 'C37',
                'pfe_value': 0.199*3,  # Pfe * 3 (trifásico)
                'pcu_cell': 'C38',
                'pcu_value': 0.626*3,  # Pcu * 3 (trifásico)
                'pr_cell': 'C47'
            },
            {
                'area': 'Preparatoria',
                'consumo_mes_cell': 'K39',
                'consumo_react_cell': 'K40',
                'kvan_cell': 'I42',
                'kvan_value': 37.5*3,  # 37.5 kVA * 3 (trifásico)
                'pfe_cell': 'I37',
                'pfe_value': 0.162*3,  # Pfe * 3 (trifásico)
                'pcu_cell': 'I38',
                'pcu_value': 0.487*3,  # Pcu * 3 (trifásico)
                'pr_cell': 'I47'
            }
        ]
        
        # Obtener datos de la base de datos
        loss_data = models.TransformerLossData.objects.filter(month=month, year=year)
        loss_data_dict = {data.area: data for data in loss_data}
        
        # Llenar datos para cada área
        for area_config in areas_config:
            data = loss_data_dict.get(area_config['area'])
            if not data:
                continue
                
            # Establecer valores en las celdas
            ws[area_config['consumo_mes_cell']] = data.monthly_consumption
            ws[area_config['consumo_react_cell']] = data.reactive_consumption
            ws[area_config['kvan_cell']] = area_config['kvan_value']
            ws[area_config['pfe_cell']] = area_config['pfe_value']
            ws[area_config['pcu_cell']] = area_config['pcu_value']
            
            # Calcular y establecer Pr
            t3 = days * 24
            t1 = 24 * days
            
            # Kvar = Consumo mes / (T1 * Fp)
            # Fp = cos(atan(Consumo Reactivo / Consumo mes))
            fp = math.cos(math.atan(data.reactive_consumption / data.monthly_consumption)) if data.monthly_consumption != 0 else 0
            kvar = data.monthly_consumption / (t1 * fp) if (t1 * fp) != 0 else 0
            
            # Pr = Pfe * T3 + (Kvar/Kvan)^2 * Pcu * T1
            pr = (area_config['pfe_value'] * t3) + ((kvar/area_config['kvan_value'])**2) * area_config['pcu_value'] * t1
            ws[area_config['pr_cell']] = pr
            
            # Actualizar total si corresponde
            if area_config['pr_cell'] in ['C17', 'I17', 'O17', 'C32', 'I32', 'C47', 'I47']:
                total_cell = area_config['pr_cell'][0] + str(int(area_config['pr_cell'][1:])) + 2  # Celda 2 filas abajo
                ws[total_cell] = f"=SUM({area_config['consumo_mes_cell']}, {area_config['pr_cell']})"
        
        # Guardar el resultado en un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        print(f"Error al generar Excel B: {str(e)}")
        raise

def find_bitacora_file(month, year):
    """Busca el archivo TablaBitacora en posibles ubicaciones"""
    import os
    from django.conf import settings
    
    month_names = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    month_name = month_names.get(month, "")
    filename = f"Tabla R 1A-Bitacora {month_name} {year}.xlsx"
    
    # Posibles ubicaciones donde el usuario podría haber guardado el archivo
    possible_locations = [
        os.path.join(settings.BASE_DIR, 'uploads'),  # Carpeta de uploads
        os.path.join(os.path.expanduser('~'), 'Downloads'),  # Descargas del usuario
        os.path.join(os.path.expanduser('~'), 'Desktop'),    # Escritorio del usuario
        os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'generated_files')  # Carpeta de archivos generados
    ]
    
    for location in possible_locations:
        file_path = os.path.join(location, filename)
        if os.path.exists(file_path):
            return file_path
    
    return None

def days_in_month(month, year):
    from calendar import monthrange
    return monthrange(year, month)[1]