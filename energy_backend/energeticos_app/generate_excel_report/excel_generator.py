# import openpyxl
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter
# from datetime import datetime
# from .models import EnergyConsumption

# YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# HEADER_FONT = Font(bold=True)
# CENTER_ALIGN = Alignment(horizontal='center')
# BORDER = Border(
#     left=Side(style='thin'),
#     right=Side(style='thin'),
#     top=Side(style='thin'),
#     bottom=Side(style='thin')
# )

# def _bitacora(energy_consumption, plan_mes=None):
    
#     try:
#         wb = openpyxl.Workbook()
        
#         # Configuración común para todas las hojas
#         areas = [
#             ('Cocina C', 'COCINA COMEDOR', 21000, 7.5, 2.5),
#             ('Ciencia A', 'CIENCIA ANIMAL', 26000, 3, 2.5),
#             ('Preparatoria', 'PREPARATORIA', 10000, 7.5, 2.5),
#             ('Sede Martí', 'JOSE MARTÍ', 20000, 7.5, 1.5),
#             ('A.Cultural', 'AREA CULTURAL', 26000, 5, 3),
#             ('Laboratorios Tec.', 'LABORATORIO TECNICO', 10000, 9.5, 2.5),
#             ('Sede Fajardo', 'MANUEL FAJARDO', 5430, 3, 1.5),
#         ]
        
#         for area_name, sheet_title, default_plan, factor_pico, factor_react in areas:
#             ws = wb.create_sheet(title=area_name)
#             current_plan = plan_mes if plan_mes is not None else default_plan
#             setup_sheet(ws, sheet_title, energy_consumption.month, energy_consumption.year, 
#                     current_plan, energy_consumption.perdidas, factor_pico, factor_react)
            
#             if energy_consumption.area == area_name:
#                 daily_data = energy_consumption.daily_data.all().order_by('-day')
#                 fill_daily_data(ws, daily_data, energy_consumption.month, energy_consumption.year, 
#                             factor_pico, factor_react)

#         if 'Sheet' in wb.sheetnames:
#             wb.remove(wb['Sheet'])
        
#         from io import BytesIO
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         return buffer.getvalue()
    
#     except Exception as e:
#         print(f"Error al generar Excel: {str(e)}")
#         raise  # Relanzar la excepción para manejarla en la vista

# #Cod_2
# def setup_sheet(ws, service_name, month, year, plan_mes, perdidas, factor_pico, factor_react):
#     # Configurar ancho de columnas
#     column_widths = {
#         'A': 3, 'B': 4, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8, 'H': 8,
#         'I': 8, 'J': 8, 'K': 8, 'L': 8, 'M': 8, 'N': 8, 'O': 8, 'P': 8,
#         'Q': 8, 'R': 8, 'S': 8, 'T': 8, 'U': 8, 'V': 8, 'W': 8, 'X': 8,
#         'Y': 8, 'Z': 8, 'AA': 8, 'AB': 8
#     }
    
#     for col, width in column_widths.items():
#         ws.column_dimensions[col].width = width
    
#     # Escribir encabezados
#     ws.merge_cells('G2:H2')
#     ws['G2'] = 'ANEXO No. 1.'
#     ws['G2'].font = HEADER_FONT
    
#     ws.merge_cells('G3:H3')
#     ws['G3'] = 'TABLA 1A'
#     ws['G3'].font = HEADER_FONT
    
#     # Celda fusionada con texto y fecha
#     ws['G4'] = f'Cumplimiento del plan de consumo de energía eléctrica total, mes: {datetime(year, month, 1).strftime("%Y-%m-%d 00:00:00")}'
#     ws.merge_cells('G4:N4')
#     ws['G4'].font = HEADER_FONT
#     ws['G4'].alignment = CENTER_ALIGN
    
#     ws.merge_cells('C7:E7')
#     ws['C6'] = f'SERVICIO: {service_name}'
#     ws['C6'].font = HEADER_FONT
    
#     # Configuración de fórmulas y valores
#     ws['S6'] = plan_mes
#     ws['U6'] = perdidas
#     ws['V6'] = f'=U6/{days_in_month(month, year)}'
    
#     # Plan Pico N y D
#     ws['O6'] = f'=S6*16.5%'
#     ws['Q6'] = f'=(S6*8)/100'

#     # Encabezados de columnas
#     ws['N6'] = f'Plan Pico N'
#     ws['P6'] = f'Plan Pico D'
#     ws['P6'] = f'Plan Mes'
#     ws['T6'] = f'Perdidas'
#     ws['U5'] = f'Total'
#     ws['V5'] = f'Diarias'
    
#     headers = [
#         ('B8', 'Día'),
#         ('C8', 'Lectura del contador'),
#         ('F8', 'Consumo diario kWh y kvarh'),
#         ('M8', 'Consumo diario Pico nocturno kWh'),
#         ('P8', 'Consumo diario pico Diurno kWh'),
#         ('R8', 'Consumo diario total kWh'),
#         ('U8', 'Consumo acumulado Pico kWh'),
#         ('X8', 'Consumo acumulado Total kWh'),
#         ('AA8', 'Firma Responsable'),
        
#         ('C9', 'MAD'), ('D9', 'DIA'), ('E9', 'PICO'), ('F9', 'TOTAL'),
#         ('G9', 'Pico Diurno 11 am'), ('H9', 'Pico Diurno 1 pm'), ('I9', 'REACT'),
#         ('J9', 'MAD'), ('K9', 'DIA'), ('L9', 'PICO'), ('M9', 'REACT'),
#         ('N9', 'Plan Pico'), ('O9', 'Real Pico'), ('P9', 'Plan - Real'),
#         ('Q9', 'Plan'), ('R9', 'Real'), ('S9', 'Plan - Real'),
#         ('T9', 'Plan Pico'), ('U9', 'Real Pico'), ('V9', 'Plan - Real'),
#         ('W9', 'Plan'), ('X9', 'Real'), ('Y9', 'Plan - Real'),
#         ('Z9', 'FP'),
#     ]

#     # # Fucionando encabezados de columnas
#     # ws.merge_cells('B8:B9')
#     # ws.merge_cells('C8:I9')

#     for cell, value in headers:
#         ws[cell] = value
#         ws[cell].font = HEADER_FONT
#         ws[cell].alignment = CENTER_ALIGN
#         ws[cell].border = BORDER
    
#     # Numeración de columnas
#     for col, num in zip(range(3, 28), range(1, 26)):
#         ws.cell(row=10, column=col, value=num)
#         ws.cell(row=10, column=col).font = HEADER_FONT
#         ws.cell(row=10, column=col).alignment = CENTER_ALIGN
#         ws.cell(row=10, column=col).border = BORDER

#         # ws.cell(row=11, column=col, value=num)
#         # ws.cell(row=11, column=col).font = HEADER_FONT
#         # ws.cell(row=11, column=col).alignment = CENTER_ALIGN
#         # ws.cell(row=11, column=col).border = BORDER
    
#     # Pie de página
#     ws.merge_cells('Y45:AA45')
#     ws['Y45'] = 'Aprobado:'
#     ws['Y45'].font = HEADER_FONT
    
#     ws.merge_cells('Y46:AA46')
#     ws['Y46'] = 'Ms.C. Jorge Luis García Santana'
    
#     ws.merge_cells('Y47:AA47')
#     ws['Y47'] = 'Director General'
    
#     ws.merge_cells('Y48:AA48')
#     ws['Y48'] = 'Universidad Camaguey'

# # #Cod_2_En-Uso
# def fill_daily_data(ws, daily_data, month, year, factor_pico, factor_react):
#     days = days_in_month(month, year)
#     start_row = 11

#     for i, data in enumerate(daily_data):
#         row = start_row + i
        
#         if data.day == 1:
#             # Primera fila (día 1) tiene fórmulas especiales
#             ws[f'B{row}'] = data.day
#             ws[f'C{row}'] = data.mad if data.mad else ''
#             ws[f'D{row}'] = data.dia if data.dia else ''
#             ws[f'E{row}'] = data.pico if data.pico else ''
#             ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
#             ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
#             ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
#             ws[f'I{row}'] = data.react if data.react else ''
#             ws[f'J{row}'] = f'=C{row}-C{row+1}'
#             ws[f'K{row}'] = f'=D{row}-D{row+1}'
#             ws[f'L{row}'] = f'=E{row}-E{row+1}'
#             ws[f'M{row}'] = f'=I{row}-I{row+1}'
#             ws[f'N{row}'] = f'=O6/{days}'
#             ws[f'O{row}'] = f'=L{row}'
#             ws[f'P{row}'] = f'=N{row}-O{row}'
#             ws[f'Q{row}'] = f'=H{row}-G{row}'
#             ws[f'R{row}'] = f'=S6/{days}'
#             ws[f'S{row}'] = f'=J{row}+K{row}+L{row}+V6'
#             ws[f'T{row}'] = f'=R{row}-S{row}'
#             ws[f'U{row}'] = f'=N{row}'
#             ws[f'V{row}'] = f'=O{row}'
#             ws[f'W{row}'] = f'=U{row}-V{row}'
#             ws[f'X{row}'] = f'=R{row}'
#             ws[f'Y{row}'] = f'=S{row}'
#             ws[f'Z{row}'] = f'=X{row}-Y{row}'
#             ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
#         else:
#             # Filas normales
#             ws[f'B{row}'] = data.day
#             ws[f'C{row}'] = data.mad if data.mad else ''
#             ws[f'D{row}'] = data.dia if data.dia else ''
#             ws[f'E{row}'] = data.pico if data.pico else ''
#             ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
#             ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
#             ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
#             ws[f'I{row}'] = data.react if data.react else ''
#             ws[f'J{row}'] = f'=C{row}-C{row+1}'
#             ws[f'K{row}'] = f'=D{row}-D{row+1}'
#             ws[f'L{row}'] = f'=E{row}-E{row+1}'
#             ws[f'M{row}'] = f'=I{row}-I{row+1}'
#             ws[f'N{row}'] = f'=N{row+1}'
#             ws[f'O{row}'] = f'=L{row}'
#             ws[f'P{row}'] = f'=N{row}-O{row}'
#             ws[f'Q{row}'] = f'=H{row}-G{row}'
#             ws[f'R{row}'] = f'=R{row+1}'
#             ws[f'S{row}'] = f'=J{row}+K{row}+L{row}'
#             ws[f'T{row}'] = f'=R{row}-S{row}'
#             ws[f'U{row}'] = f'=U{row+1}+N{row}'
#             ws[f'V{row}'] = f'=V{row+1}+O{row}'
#             ws[f'W{row}'] = f'=U{row}-V{row}'
#             ws[f'X{row}'] = f'=X{row+1}+R{row}'
#             ws[f'Y{row}'] = f'=Y{row+1}+S{row}'
#             ws[f'Z{row}'] = f'=X{row}-Y{row}'
#             ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
        
#         # Resaltar celdas importantes
#         for col in ['N', 'R', 'S']:
#             ws[f'{col}{row}'].fill = YELLOW_FILL
    
#     # Fila de cierre del mes anterior
#     cierre_row = start_row + days
#     last_day_data = daily_data.first()  # El primer registro es el día más alto (orden descendente)
    
#     ws[f'B{cierre_row}'] = 'cierre mes anterior'
#     ws[f'C{cierre_row}'] = last_day_data.mad_cierre if last_day_data.mad_cierre else ''
#     ws[f'D{cierre_row}'] = last_day_data.dia_cierre if last_day_data.dia_cierre else ''
#     ws[f'E{cierre_row}'] = last_day_data.pico_cierre if last_day_data.pico_cierre else ''
#     ws[f'F{cierre_row}'] = last_day_data.total_cierre if last_day_data.total_cierre else f'=C{cierre_row}+D{cierre_row}+E{cierre_row}'
#     ws[f'G{cierre_row}'] = last_day_data.pico_diurno_11am_cierre if last_day_data.pico_diurno_11am_cierre else ''
#     ws[f'H{cierre_row}'] = last_day_data.pico_diurno_1pm_cierre if last_day_data.pico_diurno_1pm_cierre else ''
#     ws[f'I{cierre_row}'] = last_day_data.react_cierre if last_day_data.react_cierre else ''
#     ws[f'M{cierre_row}'] = f'=SUM(M{start_row}:M{cierre_row-1})'

#     for merged_range in list(ws.merged_cells.ranges):
#         if f'AA{cierre_row}' in merged_range:
#             ws.unmerge_cells(str(merged_range))

#     ws[f'AA{cierre_row}'] = f'=AVERAGE(AA{start_row}:AA{cierre_row-1})'

# def days_in_month(month, year):
#     from calendar import monthrange
#     return monthrange(year, month)[1]

# Prueba del dia 21/04/25 - codigo funcional
# import openpyxl
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter
# from datetime import datetime
# from .models import EnergyConsumption

# YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# HEADER_FONT = Font(bold=True)
# CENTER_ALIGN = Alignment(horizontal='center')
# BORDER = Border(
#     left=Side(style='thin'),
#     right=Side(style='thin'),
#     top=Side(style='thin'),
#     bottom=Side(style='thin')
# )

# # def _bitacora(energy_consumption, plan_mes=None):
# #     try:
# #         wb = openpyxl.Workbook()
        
# #         # Configuración común para todas las hojas
# #         areas = [
# #             ('Cocina C', 'COCINA COMEDOR', 21000, 7.5, 2.5),
# #             ('Ciencia A', 'CIENCIA ANIMAL', 26000, 3, 2.5),
# #             ('Preparatoria', 'PREPARATORIA', 10000, 7.5, 2.5),
# #             ('Sede Martí', 'JOSE MARTÍ', 20000, 7.5, 1.5),
# #             ('A.Cultural', 'AREA CULTURAL', 26000, 5, 3),
# #             ('Laboratorios Tec.', 'LABORATORIO TECNICO', 10000, 9.5, 2.5),
# #             ('Sede Fajardo', 'MANUEL FAJARDO', 5430, 3, 1.5),
# #         ]

# #         for area_name, sheet_title, default_plan, factor_pico, factor_react in areas:
# #             ws = wb.create_sheet(title=area_name)
# #             current_plan = plan_mes if plan_mes is not None else default_plan
# #             setup_sheet(ws, sheet_title, energy_consumption.month, energy_consumption.year, 
# #                        current_plan, energy_consumption.perdidas, factor_pico, factor_react)
            
# #             if energy_consumption.area == area_name:
# #                 daily_data = energy_consumption.daily_data.all().order_by('day')  # Cambiado a orden ascendente
# #                 fill_daily_data(ws, daily_data, energy_consumption.month, energy_consumption.year, 
# #                               factor_pico, factor_react)
                
# #                 # Aplicar bordes a toda la tabla con datos
# #                 apply_table_borders(ws, energy_consumption.month, energy_consumption.year)

# #         if 'Sheet' in wb.sheetnames:
# #             wb.remove(wb['Sheet'])

# #         from io import BytesIO
# #         buffer = BytesIO()
# #         wb.save(buffer)
# #         buffer.seek(0)
# #         return buffer.getvalue()
# #     except Exception as e:
# #         print(f"Error al generar Excel: {str(e)}")
# #         raise

# def _bitacora(energy_consumption, plan_mes=None):
#     try:
#         wb = openpyxl.Workbook()
        
#         # Configuración común para todas las hojas
#         areas = [
#             ('Cocina C', 'COCINA COMEDOR', 21000, 7.5, 2.5),
#             ('Ciencia A', 'CIENCIA ANIMAL', 26000, 3, 2.5),
#             ('Preparatoria', 'PREPARATORIA', 10000, 7.5, 2.5),
#             ('Sede Martí', 'JOSE MARTÍ', 20000, 7.5, 1.5),
#             ('A.Cultural', 'AREA CULTURAL', 26000, 5, 3),
#             ('Laboratorios Tec.', 'LABORATORIO TECNICO', 10000, 9.5, 2.5),
#             ('Sede Fajardo', 'MANUEL FAJARDO', 5430, 3, 1.5),
#         ]
        
#         for area_name, sheet_title, default_plan, factor_pico, factor_react in areas:
#             ws = wb.create_sheet(title=area_name)
#             current_plan = plan_mes if plan_mes is not None else default_plan
#             setup_sheet(ws, sheet_title, energy_consumption.month, energy_consumption.year, 
#                        current_plan, energy_consumption.perdidas, factor_pico, factor_react)
            
#             # Obtener los datos de EnergyConsumption para esta área, mes y año
#             try:
#                 area_consumption = EnergyConsumption.objects.get(
#                     area=area_name,
#                     month=energy_consumption.month,
#                     year=energy_consumption.year
#                 )
#                 daily_data = area_consumption.daily_data.all().order_by('day')
#                 fill_daily_data(ws, daily_data, energy_consumption.month, 
#                               energy_consumption.year, factor_pico, factor_react)
#             except EnergyConsumption.DoesNotExist:
#                 # Si no hay datos para esta área, dejamos la hoja vacía
#                 pass
            
#             # Aplicar bordes a toda la tabla con datos
#             apply_table_borders(ws, energy_consumption.month, energy_consumption.year)

#         if 'Sheet' in wb.sheetnames:
#             wb.remove(wb['Sheet'])

#         from io import BytesIO
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)
#         return buffer.getvalue()
        
#     except Exception as e:
#         print(f"Error al generar Excel: {str(e)}")
#         raise


# def apply_table_borders(ws, month, year):
#     days = days_in_month(month, year)
#     first_row = 41 - days + 1  # Primera fila con datos (día 1)
#     last_row = 42  # Fila de cierre
    
#     # Definir el rango de columnas con datos (de B a AB)
#     for row in range(first_row, last_row + 1):
#         for col in range(2, 29):  # Columnas B (2) a AB (28)
#             cell = ws.cell(row=row, column=col)
#             cell.border = BORDER

# def setup_sheet(ws, service_name, month, year, plan_mes, perdidas, factor_pico, factor_react):
#     # Configurar ancho de columnas
#     column_widths = {
#         'A': 10, 'B': 10, 'C': 25, 'D': 25, 'E': 25, 'F': 25, 'G': 25, 'H': 25, 'I': 25,
#         'J': 25, 'K': 25, 'L': 25, 'M': 25, 'N': 25, 'O': 25, 'P': 25, 'Q': 25, 'R': 25,
#         'S': 25, 'T': 25, 'U': 25, 'V': 25, 'W': 25, 'X': 25, 'Y': 25, 'Z': 25, 'AA': 25, 'AB': 25
#     }
#     for col, width in column_widths.items():
#         ws.column_dimensions[col].width = width

#     # Escribir encabezados
#     # Modificado para evitar el error de celdas fusionadas
#     cell_g2 = ws['G2']
#     ws.merge_cells('G2:H2')
#     cell_g2.value = 'ANEXO No. 1.'
#     cell_g2.font = HEADER_FONT
    
#     cell_g3 = ws['G3']
#     ws.merge_cells('G3:H3')
#     cell_g3.value = 'TABLA 1A'
#     cell_g3.font = HEADER_FONT

#     # Celda fusionada con texto y fecha
#     cell_g4 = ws['G4']
#     ws.merge_cells('G4:M4')
#     cell_g4.value = f'Cumplimiento del plan de consumo de energía eléctrica total, mes: {datetime(year, month, 1).strftime("%Y-%m-%d 00:00:00")}'
#     cell_g4.font = HEADER_FONT
#     cell_g4.alignment = CENTER_ALIGN

#     ws.merge_cells('C7:E7')
#     cell_c6 = ws['C6']
#     cell_c6.value = f'SERVICIO: {service_name}'
#     cell_c6.font = HEADER_FONT

#     # Configuración de fórmulas y valores
#     ws['S6'] = plan_mes
#     ws['U6'] = perdidas
#     ws['V6'] = f'=U6/{days_in_month(month, year)}'

#     # Plan Pico N y D
#     ws['O6'] = f'=S6*16.5%'
#     ws['Q6'] = f'=(S6*8)/100'

#     # Encabezados de columnas
#     ws['N6'] = 'Plan Pico N'
#     ws['P6'] = 'Plan Pico D'
#     ws['P6'] = 'Plan Mes'
#     ws['T6'] = 'Perdidas'
#     ws['U5'] = 'Total'
#     ws['V5'] = 'Diarias'

#     headers = [
#         ('B8', 'Día'), ('C8', 'Lectura del contador'), ('F8', 'Consumo diario kWh y kvarh'),
#         ('M8', 'Consumo diario Pico nocturno kWh'), ('P8', 'Consumo diario pico Diurno kWh'),
#         ('R8', 'Consumo diario total kWh'), ('U8', 'Consumo acumulado Pico kWh'),
#         ('X8', 'Consumo acumulado Total kWh'), ('AA8', 'Firma Responsable'),
#         ('C9', 'MAD'), ('D9', 'DIA'), ('E9', 'PICO'), ('F9', 'TOTAL'),
#         ('G9', 'Pico Diurno 11 am'), ('H9', 'Pico Diurno 1 pm'), ('I9', 'REACT'),
#         ('J9', 'MAD'), ('K9', 'DIA'), ('L9', 'PICO'), ('M9', 'REACT'),
#         ('N9', 'Plan Pico'), ('O9', 'Real Pico'), ('P9', 'Plan - Real'),
#         ('Q9', 'Plan'), ('R9', 'Real'), ('S9', 'Plan - Real'), ('T9', 'Plan Pico'),
#         ('U9', 'Real Pico'), ('V9', 'Plan - Real'), ('W9', 'Plan'), ('X9', 'Real'),
#         ('Y9', 'Plan - Real'), ('Z9', 'FP'),
#     ]
    
#     for cell, value in headers:
#         ws[cell] = value
#         ws[cell].font = HEADER_FONT
#         ws[cell].alignment = CENTER_ALIGN
#         ws[cell].border = BORDER

#     # # Numeración de columnas
#     # for col, num in zip(range(3, 28), range(1, 26)):
#     #     ws.cell(row=10, column=col, value=num)
#     #     ws.cell(row=10, column=col).font = HEADER_FONT
#     #     ws.cell(row=10, column=col).alignment = CENTER_ALIGN
#     #     ws.cell(row=10, column=col).border = BORDER

#     # Numeración de columnas (ajustado para que F10 esté vacío y G10 empiece con 4)
#     column_numbers = [
#         ('C10', 1), ('D10', 2), ('E10', 3), 
#         ('G10', 4), ('H10', 5), ('I10', 6),  # F10 se salta, empieza en G10 con 4
#         ('J10', 7), ('K10', 8), ('L10', 9), ('M10', 10),
#         ('N10', 11), ('O10', 12), ('P10', 13),
#         ('Q10', 14), ('R10', 15), ('S10', 16),
#         ('T10', 17), ('U10', 18), ('V10', 19),
#         ('W10', 20), ('X10', 21), ('Y10', 22),
#         ('Z10', 23), ('AA10', 24), ('AB10', 25)
#     ]
    
#     for cell, num in column_numbers:
#         ws[cell] = num
#         ws[cell].font = HEADER_FONT
#         ws[cell].alignment = CENTER_ALIGN
#         ws[cell].border = BORDER


#     # Pie de página
#     ws.merge_cells('Y45:AA45')
#     ws['Y45'] = 'Aprobado:'
#     ws['Y45'].font = HEADER_FONT
    
#     ws.merge_cells('Y46:AA46')
#     ws['Y46'] = 'Ms.C. Jorge Luis García Santana'
    
#     ws.merge_cells('Y47:AA47')
#     ws['Y47'] = 'Director General'
    
#     ws.merge_cells('Y48:AA48')
#     ws['Y48'] = 'Universidad Camaguey'

# def fill_daily_data(ws, daily_data, month, year, factor_pico, factor_react):
#     days = days_in_month(month, year)
#     base_row = 41  # Fila base donde empezará el día 1
    
#     for data in daily_data:
#         row = base_row - data.day + 1  # Calcula la fila basada en el día
        
#         if data.day == 1:
#             # Primera fila (día 1) tiene fórmulas especiales
#             ws[f'B{row}'] = data.day
#             ws[f'C{row}'] = data.mad if data.mad else ''
#             ws[f'D{row}'] = data.dia if data.dia else ''
#             ws[f'E{row}'] = data.pico if data.pico else ''
#             ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
#             ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
#             ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
#             ws[f'I{row}'] = data.react if data.react else ''
#             ws[f'J{row}'] = f'=C{row}-C{row+1}'
#             ws[f'K{row}'] = f'=D{row}-D{row+1}'
#             ws[f'L{row}'] = f'=E{row}-E{row+1}'
#             ws[f'M{row}'] = f'=I{row}-I{row+1}'
#             ws[f'N{row}'] = f'=O6/{days}'
#             ws[f'O{row}'] = f'=L{row}'
#             ws[f'P{row}'] = f'=N{row}-O{row}'
#             ws[f'Q{row}'] = f'=H{row}-G{row}'
#             ws[f'R{row}'] = f'=S6/{days}'
#             ws[f'S{row}'] = f'=J{row}+K{row}+L{row}+V6'
#             ws[f'T{row}'] = f'=R{row}-S{row}'
#             ws[f'U{row}'] = f'=N{row}'
#             ws[f'V{row}'] = f'=O{row}'
#             ws[f'W{row}'] = f'=U{row}-V{row}'
#             ws[f'X{row}'] = f'=R{row}'
#             ws[f'Y{row}'] = f'=S{row}'
#             ws[f'Z{row}'] = f'=X{row}-Y{row}'
#             ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
#         else:
#             # Filas normales
#             ws[f'B{row}'] = data.day
#             ws[f'C{row}'] = data.mad if data.mad else ''
#             ws[f'D{row}'] = data.dia if data.dia else ''
#             ws[f'E{row}'] = data.pico if data.pico else ''
#             ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
#             ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
#             ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
#             ws[f'I{row}'] = data.react if data.react else ''
#             ws[f'J{row}'] = f'=C{row}-C{row+1}'
#             ws[f'K{row}'] = f'=D{row}-D{row+1}'
#             ws[f'L{row}'] = f'=E{row}-E{row+1}'
#             ws[f'M{row}'] = f'=I{row}-I{row+1}'
#             ws[f'N{row}'] = f'=N{row+1}'
#             ws[f'O{row}'] = f'=L{row}'
#             ws[f'P{row}'] = f'=N{row}-O{row}'
#             ws[f'Q{row}'] = f'=H{row}-G{row}'
#             ws[f'R{row}'] = f'=R{row+1}'
#             ws[f'S{row}'] = f'=J{row}+K{row}+L{row}'
#             ws[f'T{row}'] = f'=R{row}-S{row}'
#             ws[f'U{row}'] = f'=U{row+1}+N{row}'
#             ws[f'V{row}'] = f'=V{row+1}+O{row}'
#             ws[f'W{row}'] = f'=U{row}-V{row}'
#             ws[f'X{row}'] = f'=X{row+1}+R{row}'
#             ws[f'Y{row}'] = f'=Y{row+1}+S{row}'
#             ws[f'Z{row}'] = f'=X{row}-Y{row}'
#             ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'

#         # Resaltar celdas importantes
#         for col in ['N', 'R', 'S']:
#             ws[f'{col}{row}'].fill = YELLOW_FILL

#     # Fila de cierre del mes anterior (ahora en la fila 11)
#     cierre_row = 42
    
#     # last_day_data = daily_data.last()  # Ahora usamos last() porque los datos están en orden ascendente
    
#     # ws[f'B{cierre_row}'] = 'cierre mes anterior'
#     # ws[f'C{cierre_row}'] = last_day_data.mad_cierre if last_day_data.mad_cierre else ''
#     # ws[f'D{cierre_row}'] = last_day_data.dia_cierre if last_day_data.dia_cierre else ''
#     # ws[f'E{cierre_row}'] = last_day_data.pico_cierre if last_day_data.pico_cierre else ''
#     # ws[f'F{cierre_row}'] = last_day_data.total_cierre if last_day_data.total_cierre else f'=C{cierre_row}+D{cierre_row}+E{cierre_row}'
#     # ws[f'G{cierre_row}'] = last_day_data.pico_diurno_11am_cierre if last_day_data.pico_diurno_11am_cierre else ''
#     # ws[f'H{cierre_row}'] = last_day_data.pico_diurno_1pm_cierre if last_day_data.pico_diurno_1pm_cierre else ''
#     # ws[f'I{cierre_row}'] = last_day_data.react_cierre if last_day_data.react_cierre else ''
   

#     if daily_data.exists():
#         last_day_data = daily_data.last()  # Ahora usamos last() porque los datos están en orden ascendente
        
#         ws[f'B{cierre_row}'] = 'cierre mes anterior'
#         ws[f'C{cierre_row}'] = last_day_data.mad_cierre if last_day_data.mad_cierre else ''
#         ws[f'D{cierre_row}'] = last_day_data.dia_cierre if last_day_data.dia_cierre else ''
#         ws[f'E{cierre_row}'] = last_day_data.pico_cierre if last_day_data.pico_cierre else ''
#         ws[f'F{cierre_row}'] = last_day_data.total_cierre if last_day_data.total_cierre else f'=C{cierre_row}+D{cierre_row}+E{cierre_row}'
#         ws[f'G{cierre_row}'] = last_day_data.pico_diurno_11am_cierre if last_day_data.pico_diurno_11am_cierre else ''
#         ws[f'H{cierre_row}'] = last_day_data.pico_diurno_1pm_cierre if last_day_data.pico_diurno_1pm_cierre else ''
#         ws[f'I{cierre_row}'] = last_day_data.react_cierre if last_day_data.react_cierre else ''
#     else:
#         #si no hay datos, deja las celdas vacias
#         ws[f'B{cierre_row}'] = 'cierre mes anterior'
#         ws[f'C{cierre_row}'] = ''
#         ws[f'D{cierre_row}'] = ''
#         ws[f'E{cierre_row}'] = ''
#         ws[f'F{cierre_row}'] = ''
#         ws[f'G{cierre_row}'] = ''
#         ws[f'H{cierre_row}'] = ''
#         ws[f'I{cierre_row}'] = ''


#     # Calcular el rango de filas con datos (de 41 hacia arriba hasta el día 1)
#     first_data_row = base_row - days + 1
#     ws[f'M{cierre_row}'] = f'=SUM(M{first_data_row}:M{base_row})'
    
#     # Manejar celdas fusionadas para el promedio
#     for merged_range in list(ws.merged_cells.ranges):
#         if f'AA{cierre_row}' in str(merged_range):
#             ws.unmerge_cells(str(merged_range))
    
#     ws[f'AA{cierre_row}'] = f'=AVERAGE(AA{first_data_row}:AA{base_row})'

# def days_in_month(month, year):
#     from calendar import monthrange
#     return monthrange(year, month)[1]


#Cargando desde el modeloA - Nueva prueba/funcional

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from django.conf import settings
from .models import EnergyConsumption

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center')
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# def load_template():
#     """Carga el archivo modeloA.xlsx desde el sistema de archivos"""
#     template_path = os.path.join(settings.BASE_DIR, 'templates', 'modeloA.xlsx')
#     return openpyxl.load_workbook(template_path)

def load_template():
       """Carga el archivo modeloA.xlsx desde los recursos de la energeticos_app"""
       import os
       from django.conf import settings
       
       # Ruta absoluta al archivo
       template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloA.xlsx')
       
       if not os.path.exists(template_path):
           raise FileNotFoundError(
               f"No se encontró el archivo modelo en: {template_path}"
           )
       
       return openpyxl.load_workbook(template_path)

def generate_excel_report_bitacora(energy_consumption, plan_mes=None):
    try:
        # Cargar el modelo en lugar de crear un nuevo libro
        wb = load_template()
        
        # Configuración común para todas las hojas
        areas_config = [
            ('Cocina C', 'COCINA COMEDOR', 21000, 7.5, 2.5),
            ('Ciencia A', 'CIENCIA ANIMAL', 26000, 3, 2.5),
            ('Preparatoria', 'PREPARATORIA', 10000, 7.5, 2.5),
            ('Sede Martí', 'JOSE MARTÍ', 20000, 7.5, 1.5),
            ('A.Cultural', 'AREA CULTURAL', 26000, 5, 3),
            ('Laboratorios Tec.', 'LABORATORIO TECNICO', 10000, 9.5, 2.5),
            ('Sede Fajardo', 'MANUEL FAJARDO', 5430, 3, 1.5),
        ]

        for area_name, sheet_title, default_plan, factor_pico, factor_react in areas_config:
            # Verificar si la hoja existe en el modelo
            if area_name not in wb.sheetnames:
                continue
                
            ws = wb[area_name]
            current_plan = plan_mes if plan_mes is not None else default_plan
            
            # Actualizar la hoja con los datos
            update_sheet(
                ws, sheet_title, energy_consumption.month, 
                energy_consumption.year, current_plan, 
                energy_consumption.perdidas, factor_pico, factor_react
            )
            
            # Obtener los datos para esta área
            try:
                area_consumption = EnergyConsumption.objects.get(
                    area=area_name,
                    month=energy_consumption.month,
                    year=energy_consumption.year
                )
                daily_data = area_consumption.daily_data.all().order_by('day')
                fill_daily_data(
                    ws, daily_data, energy_consumption.month, 
                    energy_consumption.year, factor_pico, factor_react
                )
            except EnergyConsumption.DoesNotExist:
                # Si no hay datos para esta área, dejamos la hoja como está
                pass

        # Guardar el resultado en un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        print(f"Error al generar Excel: {str(e)}")
        raise

def update_sheet(ws, service_name, month, year, plan_mes, perdidas, factor_pico, factor_react):
    """Actualiza una hoja del modelo con los datos específicos"""
    # Actualizar valores básicos
    ws['S6'] = plan_mes
    ws['U6'] = perdidas
    ws['V6'] = f'=U6/{days_in_month(month, year)}'
    
    # Actualizar fórmulas de Plan Pico N y D
    ws['O6'] = f'=S6*16.5%'
    ws['Q6'] = f'=(S6*8)/100'
    
    # Nombre del mes en español
    month_names = {
        1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
        5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
        9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
    }
    month_name = month_names.get(month, "")
    
    # Actualizar fecha en el encabezado
    # ws['G4'] = f'Cumplimiento del plan de consumo de energía eléctrica total, mes: {datetime(year, month, 1).strftime("%Y-%m-%d 00:00:00")}'
    ws['N4'] = f'{month_name}-{year}'

    # Actualizar nombre del servicio
    ws['C6'] = f'SERVICIO: {service_name}'

def fill_daily_data(ws, daily_data, month, year, factor_pico, factor_react):
    """Llena los datos diarios en la hoja de cálculo"""
    days = days_in_month(month, year)
    base_row = 41  # Fila base donde empezará el día 1
    
    for data in daily_data:
        row = base_row - data.day + 1  # Calcula la fila basada en el día
        
        # Escribir valores básicos
        ws[f'B{row}'] = data.day
        ws[f'C{row}'] = data.mad if data.mad else ''
        ws[f'D{row}'] = data.dia if data.dia else ''
        ws[f'E{row}'] = data.pico if data.pico else ''
        ws[f'I{row}'] = data.react if data.react else ''
        
        # Configurar fórmulas según el día
        if data.day == 1:
            # Fórmulas especiales para el día 1
            ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
            ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
            ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
            ws[f'J{row}'] = f'=C{row}-C{row+1}'
            ws[f'K{row}'] = f'=D{row}-D{row+1}'
            ws[f'L{row}'] = f'=E{row}-E{row+1}'
            ws[f'M{row}'] = f'=I{row}-I{row+1}'
            ws[f'N{row}'] = f'=O6/{days}'
            ws[f'O{row}'] = f'=L{row}'
            ws[f'P{row}'] = f'=N{row}-O{row}'
            ws[f'Q{row}'] = f'=H{row}-G{row}'
            ws[f'R{row}'] = f'=S6/{days}'
            ws[f'S{row}'] = f'=J{row}+K{row}+L{row}+V6'
            ws[f'T{row}'] = f'=R{row}-S{row}'
            ws[f'U{row}'] = f'=N{row}'
            ws[f'V{row}'] = f'=O{row}'
            ws[f'W{row}'] = f'=U{row}-V{row}'
            ws[f'X{row}'] = f'=R{row}'
            ws[f'Y{row}'] = f'=S{row}'
            ws[f'Z{row}'] = f'=X{row}-Y{row}'
            ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
        else:
            # Fórmulas para días normales
            ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
            ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
            ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
            ws[f'J{row}'] = f'=C{row}-C{row+1}'
            ws[f'K{row}'] = f'=D{row}-D{row+1}'
            ws[f'L{row}'] = f'=E{row}-E{row+1}'
            ws[f'M{row}'] = f'=I{row}-I{row+1}'
            ws[f'N{row}'] = f'=N{row+1}'
            ws[f'O{row}'] = f'=L{row}'
            ws[f'P{row}'] = f'=N{row}-O{row}'
            ws[f'Q{row}'] = f'=H{row}-G{row}'
            ws[f'R{row}'] = f'=R{row+1}'
            ws[f'S{row}'] = f'=J{row}+K{row}+L{row}'
            ws[f'T{row}'] = f'=R{row}-S{row}'
            ws[f'U{row}'] = f'=U{row+1}+N{row}'
            ws[f'V{row}'] = f'=V{row+1}+O{row}'
            ws[f'W{row}'] = f'=U{row}-V{row}'
            ws[f'X{row}'] = f'=X{row+1}+R{row}'
            ws[f'Y{row}'] = f'=Y{row+1}+S{row}'
            ws[f'Z{row}'] = f'=X{row}-Y{row}'
            ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
        
        # Resaltar celdas importantes
        for col in ['N41', 'R41']:
            ws[f'{col}{row}'].fill = YELLOW_FILL
    
    # Fila de cierre del mes anterior
    cierre_row = 42
    if daily_data.exists():
        last_day_data = daily_data.last()
        ws[f'B{cierre_row}'] = 'cierre mes anterior'
        ws[f'C{cierre_row}'] = last_day_data.mad_cierre if last_day_data.mad_cierre else ''
        ws[f'D{cierre_row}'] = last_day_data.dia_cierre if last_day_data.dia_cierre else ''
        ws[f'E{cierre_row}'] = last_day_data.pico_cierre if last_day_data.pico_cierre else ''
        ws[f'F{cierre_row}'] = last_day_data.total_cierre if last_day_data.total_cierre else f'=C{cierre_row}+D{cierre_row}+E{cierre_row}'
        ws[f'G{cierre_row}'] = last_day_data.pico_diurno_11am_cierre if last_day_data.pico_diurno_11am_cierre else ''
        ws[f'H{cierre_row}'] = last_day_data.pico_diurno_1pm_cierre if last_day_data.pico_diurno_1pm_cierre else ''
        ws[f'I{cierre_row}'] = last_day_data.react_cierre if last_day_data.react_cierre else ''
    else:
        ws[f'B{cierre_row}'] = 'cierre mes anterior'
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{cierre_row}'] = ''

    # Calcular el rango de filas con datos
    first_data_row = base_row - days + 1
    ws[f'M{cierre_row}'] = f'=SUM(M{first_data_row}:M{base_row})'
    ws[f'AA{cierre_row}'] = f'=AVERAGE(AA{first_data_row}:AA{base_row})'

def days_in_month(month, year):
    from calendar import monthrange
    return monthrange(year, month)[1]