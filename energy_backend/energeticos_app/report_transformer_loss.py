import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from django.conf import settings
from calendar import monthrange
import math
from .models import EnergyConsumption
from . import models

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center')
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_template_b():
    """Carga el archivo modeloB.xlsx desde los recursos de la energeticos_app"""
    # import os
    # from django.conf import settings
    
    # Ruta absoluta al archivo
    template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloB.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró el archivo modeloB en: {template_path}"
        )
    return openpyxl.load_workbook(template_path)


def generate_excel_report_transformer_loss(month, year):
    """Genera el reporte de pérdidas de transformación (modeloB.xlsx)"""
    try:
        wb = load_template_b()
        ws = wb['Pr']
        
        # Configurar mes y año
        month_names = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
            7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
        month_name = month_names.get(int(month), "")
        ws['B3'] = f"Cálculo de las pérdidas de transformación eléctrica. Mes {month_name} {int(year)}"
        
        # Configurar días y horas
        days = float(days_in_month(int(month), int(year)))
        hours = 24
        
        # Configurar días en todas las celdas necesarias
        days_cells = ['E7', 'F7', 'K7', 'L7', 'Q7', 'R7', 'E22', 'F22', 'K22', 'L22', 'E37', 'F37', 'K37', 'L37']
        for cell in days_cells:
            ws[cell] = days
        
        # Configurar horas
        hours_cells = ['E8', 'K8', 'Q8', 'E23', 'K23', 'E38', 'K38']
        for cell in hours_cells:
            ws[cell] = hours
        
        # Obtener datos de la base de datos
        loss_data = models.TransformerLossData.objects.filter(month=int(month), year=int(year))
        
        # Configuración dinámica de áreas basada en los datos de la base de datos
        areas_config = []
        position_configs = [
            # Configuración para las primeras 3 áreas
            {
                'area_name_cell': 'B4',
                'consumo_mes_cell': 'E9',
                'consumo_react_cell': 'E10',
                'kvan_cell': 'C12',
                'pfe_cell': 'C7',
                'pcu_cell': 'C8',
                't3_cell': 'C9',
                't1_cell': 'C10',
                'fp_cell': 'C11',
                'kvar_cell': 'C15',
                'pr_cell': 'C17',
                'total_cell': 'F17'
            },
            {
                'area_name_cell': 'H4',
                'consumo_mes_cell': 'K9',
                'consumo_react_cell': 'K10',
                'kvan_cell': 'I12',
                'pfe_cell': 'I7',
                'pcu_cell': 'I8',
                't3_cell': 'I9',
                't1_cell': 'I10',
                'fp_cell': 'I11',
                'kvar_cell': 'I15',
                'pr_cell': 'I17',
                'total_cell': 'L17'
            },
            {
                'area_name_cell': 'N4',
                'consumo_mes_cell': 'Q9',
                'consumo_react_cell': 'Q10',
                'kvan_cell': 'O12',
                'pfe_cell': 'O7',
                'pcu_cell': 'O8',
                't3_cell': 'O9',
                't1_cell': 'O10',
                'fp_cell': 'O11',
                'kvar_cell': 'O15',
                'pr_cell': 'O17',
                'total_cell': 'R17'
            },
            # Configuración para las siguientes 3 áreas
            {
                'area_name_cell': 'B19',
                'consumo_mes_cell': 'E24',
                'consumo_react_cell': 'E25',
                'kvan_cell': 'C27',
                'pfe_cell': 'C22',
                'pcu_cell': 'C23',
                't3_cell': 'C24',
                't1_cell': 'C25',
                'fp_cell': 'C26',
                'kvar_cell': 'C30',
                'pr_cell': 'C32',
                'total_cell': 'F32'
            },
            {
                'area_name_cell': 'H19',
                'consumo_mes_cell': 'K24',
                'consumo_react_cell': 'K25',
                'kvan_cell': 'I27',
                'pfe_cell': 'I22',
                'pcu_cell': 'I23',
                't3_cell': 'I24',
                't1_cell': 'I25',
                'fp_cell': 'I26',
                'kvar_cell': 'I30',
                'pr_cell': 'I32',
                'total_cell': 'L32'
            },
            {
                'area_name_cell': 'B34',
                'consumo_mes_cell': 'E39',
                'consumo_react_cell': 'E40',
                'kvan_cell': 'C42',
                'pfe_cell': 'C37',
                'pcu_cell': 'C38',
                't3_cell': 'C39',
                't1_cell': 'C40',
                'fp_cell': 'C41',
                'kvar_cell': 'C45',
                'pr_cell': 'C47',
                'total_cell': 'F47'
            },
            {
                'area_name_cell': 'H34',
                'consumo_mes_cell': 'K39',
                'consumo_react_cell': 'K40',
                'kvan_cell': 'I42',
                'pfe_cell': 'I37',
                'pcu_cell': 'I38',
                't3_cell': 'I39',
                't1_cell': 'I40',
                'fp_cell': 'I41',
                'kvar_cell': 'I45',
                'pr_cell': 'I47',
                'total_cell': 'L47'
            }
        ]
        
        # Crear configuración dinámica para cada área encontrada en la base de datos
        for i, data in enumerate(loss_data):
            if i >= len(position_configs):
                break  # Solo manejamos hasta 7 áreas como máximo
                
            config = position_configs[i].copy()
            config['area'] = data.area
            config['kvan_value'] = float(data.kva) * 3
            config['pfe_value'] = float(data.pfe) * 3
            config['pcu_value'] = float(data.pcu) * 3
            config['reactive_consumption'] = float(data.reactive_consumption)
            config['monthly_consumption'] = float(data.monthly_consumption)
            areas_config.append(config)
        
        # Llenar datos para cada área
        for area_config in areas_config:
            try:
                # Escribir el nombre del área
                ws[area_config['area_name_cell']] = area_config['area']
                
                # Asignar valores directos
                ws[area_config['consumo_mes_cell']] = area_config['monthly_consumption']
                ws[area_config['consumo_react_cell']] = area_config['reactive_consumption']
                ws[area_config['kvan_cell']] = float(area_config['kvan_value'])
                
                # Asignar valores de Pfe y Pcu
                ws[area_config['pfe_cell']] = float(area_config['pfe_value'])
                ws[area_config['pcu_cell']] = float(area_config['pcu_value'])
                
                # Calcular y establecer fórmulas derivadas
                ws[area_config['t3_cell']] = f"={days}*{hours}"  # T3 = días * horas
                ws[area_config['t1_cell']] = f"={hours}*{days}"  # T1 = horas * días
                ws[area_config['fp_cell']] = f"=COS(ATAN({area_config['consumo_react_cell']}/{area_config['consumo_mes_cell']}))"
                ws[area_config['kvar_cell']] = f"={area_config['consumo_mes_cell']}/({area_config['t1_cell']}*{area_config['fp_cell']})"

                # Fórmula de Pr corregida según el modelo Excel
                ws[area_config['pr_cell']] = (
                    f"={area_config['pfe_cell']}*{area_config['t3_cell']}+"
                    f"(({area_config['kvar_cell']}/{area_config['kvan_cell']})*"
                    f"({area_config['kvar_cell']}/{area_config['kvan_cell']}))*"
                    f"{area_config['pcu_cell']}*{area_config['t1_cell']}"
                )
                
                # Fórmula de total
                ws[area_config['total_cell']] = f"={area_config['consumo_mes_cell']}+{area_config['pr_cell']}"
                
            except Exception as e:
                print(f"Error procesando área {area_config.get('area', '')}: {str(e)}")
                continue
        
        # Configurar tabla de transformadores monofásicos
        transformers = [
            {'kva': 5, 'pfe': 0.046, 'pcu': 0.107},
            {'kva': 10, 'pfe': 0.065, 'pcu': 0.18},
            {'kva': 15, 'pfe': 0.084, 'pcu': 0.251},
            {'kva': 25, 'pfe': 0.115, 'pcu': 0.389},
            {'kva': 37.5, 'pfe': 0.162, 'pcu': 0.487},
            {'kva': 50, 'pfe': 0.199, 'pcu': 0.626},
            {'kva': 75, 'pfe': 0.269, 'pcu': 0.882},
            {'kva': 100, 'pfe': 0.332, 'pcu': 1.185},
            {'kva': 167, 'pfe': 0.482, 'pcu': 1.893},
            {'kva': 250, 'pfe': 0.66, 'pcu': 2.802},
            {'kva': 333, 'pfe': 0.83, 'pcu': 3.587}
        ]
        
        # Llenar tabla de transformadores (columnas S, T, U)
        start_row = 22
        for i, transformer in enumerate(transformers):
            row = start_row + i
            ws[f'S{row}'] = transformer['kva']
            ws[f'T{row}'] = transformer['pfe']
            ws[f'U{row}'] = transformer['pcu']
        
        # Guardar el resultado
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        print(f"Error al generar Excel B: {str(e)}")
        raise


def find_bitacora_file(month, year):
    """Busca el archivo TablaBitacora en posibles ubicaciones"""
    # import os
    # from django.conf import settings
    
    month_names = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    month_name = month_names.get(month, "")
    filename = f"Tabla R 1A-Bitacora {month_name} {year}.xlsx"
    
    # Posibles ubicaciones donde el usuario podría haber guardado el archivo
    possible_locations = [
        os.path.join(settings.BASE_DIR, 'uploads'),  # Carpeta de uploads
        os.path.join(os.path.expanduser('~'), 'Downloads'),  # Descargas del usuario
        os.path.join(os.path.expanduser('~'), 'Desktop'),    # Escritorio del usuario
        os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'generated_files')  # Carpeta de archivos generados
    ]
    
    for location in possible_locations:
        file_path = os.path.join(location, filename)
        if os.path.exists(file_path):
            return file_path
    
    return None

def days_in_month(month, year):
    return monthrange(year, month)[1]

