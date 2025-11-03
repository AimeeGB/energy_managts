from decimal import Decimal, InvalidOperation, getcontext
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os
from django.conf import settings
from .models import EnergyConsumption, AreaLocales, TransformerLossData, TablaBitacoraData, TablaBitacoraDataSelete, DailyConsumption
from calendar import monthrange
from django.db import transaction
import math
from django.db.models import Avg, Sum

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal='center')
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def load_template():
    """Carga el archivo modeloA.xlsx desde los recursos de la energeticos_app"""
    import os
    from django.conf import settings
    # template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloA.xlsx')
    template_path = os.path.join(settings.BASE_DIR, 'energeticos_app', 'resources', 'modeloA - 02.xlsx')
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"No se encontró el archivo modelo en: {template_path}"
        )
    return openpyxl.load_workbook(template_path)

#Codigo de funcional 02/06/25
def generate_excel_report_bitacora(energy_consumption, plan_mes=None):
    try:
        wb = load_template()
        
        # Obtener todas las áreas activas ordenadas por nombre
        areas_activas = AreaLocales.objects.filter(activo=True).order_by('nombre')

        
        # Renombrar y duplicar hojas según sea necesario
        sheet_names = [sheet.title for sheet in wb.worksheets]
        
        # Si hay más áreas activas que hojas, duplicar la última hoja hasta tener suficientes
        if len(areas_activas) > len(sheet_names):
            last_sheet = wb[sheet_names[-1]]
            for i in range(len(sheet_names), len(areas_activas)):
                new_sheet = wb.copy_worksheet(last_sheet)
                new_sheet.title = f"hoja{i+1}"
        
        # Renombrar las hojas según las áreas activas
        for i, area in enumerate(areas_activas):
            sheet_index = i if i < len(wb.sheetnames) else -1  # Usar la última hoja si hay más áreas que hojas
            sheet_name = wb.sheetnames[sheet_index]
            ws = wb[sheet_name]
            
            # Cambiar el nombre de la hoja al nombre del área
            if sheet_name != area.nombre_corto:
                try:
                    ws.title = area.nombre_corto
                except:
                    # Si hay conflicto de nombres, mantener el nombre original
                    pass
            
            try:
                area_data = EnergyConsumption.objects.get(
                    area=area,
                    month=energy_consumption.month,
                    year=energy_consumption.year
                )
                current_plan = area_data.plan_mes
                current_perdidas = area_data.perdidas

            except EnergyConsumption.DoesNotExist:
                current_plan = area.plan_default
                current_perdidas = 0
            
            update_sheet(
                ws,
                area.nombre,
                energy_consumption.month,
                energy_consumption.year,
                current_plan,
                current_perdidas,
                area.factor_pico,
                area.factor_react
            )
            
            try:
                area_consumption = EnergyConsumption.objects.get(
                    area=area,
                    month=energy_consumption.month,
                    year=energy_consumption.year
                )
                daily_data = area_consumption.daily_data.all().order_by('day')
                fill_daily_data(
                    ws,
                    daily_data,
                    energy_consumption.month,
                    energy_consumption.year,
                    area.factor_pico,
                    area.factor_react
                )
            except EnergyConsumption.DoesNotExist:
                pass

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        #Codigo de prueba 03/06/25
        # Create folder if it doesn't exist
        report_folder = os.path.join(settings.BASE_DIR, 'energeticos_app', 'generate_excel_report')
        os.makedirs(report_folder, exist_ok=True)
        
        month_names = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
            7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
        month_name = month_names.get(energy_consumption.month, "").lower()
        filename = f"Tabla R 1A-Bitacora {month_name} {energy_consumption.year}.xlsx"
        filepath = os.path.join(report_folder, filename)
        
        # Save the file
        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())

        return buffer.getvalue()
        
    except Exception as e:
        print(f"Error al generar Excel: {str(e)}")
        raise

def update_sheet(ws, service_name, month, year, plan_mes, perdidas, factor_pico, factor_react):
    """Actualiza una hoja del modelo con los datos específicos"""
    # Actualizar valores básicos
    ws['S6'] = plan_mes
    ws['U6'] = perdidas
    ws['V6'] = f'=U6/{days_in_month(month, year)}'
    
    # Actualizar fórmulas de Plan Pico N y D
    ws['O6'] = f'=S6*16.5%'
    ws['Q6'] = f'=(S6*8)/100'
    
    # Nombre del mes en español
    month_names = {
        1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
        5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
        9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
    }
    month_name = month_names.get(month, "")
    
    # Actualizar fecha en el encabezado
    # ws['G4'] = f'Cumplimiento del plan de consumo de energía eléctrica total, mes: {datetime(year, month, 1).strftime("%Y-%m-%d 00:00:00")}'
    ws['N4'] = f'{month_name}-{year}'

    # Actualizar nombre del servicio
    ws['C6'] = f'SERVICIO: {service_name}'

# Configuración de precisión decimal
getcontext().prec = 10

def calculate_and_save_direct_values(ws, daily_data, month, year, factor_pico, factor_react):
    """Calcula valores directamente y los guarda en la base de datos, incluyendo cierre del mes anterior"""
    days_in_month = monthrange(year, month)[1]
    # base_row = 41
    
    # Obtener configuración del área
    area = daily_data[0].energy_consumption.area
    plan_mes = daily_data[0].energy_consumption.plan_mes
    perdidas = daily_data[0].energy_consumption.perdidas
    
    daily_plan_pico = (plan_mes * Decimal('0.165')) / days_in_month
    daily_plan = plan_mes / days_in_month
    daily_loss = perdidas / days_in_month

    def safe_decimal(value):
        return Decimal(value) if value is not None else Decimal('0')
    
    # Ordenar los datos diarios por día de forma descendente (día 1 primero)
    # sorted_daily_data = sorted(daily_data, key=lambda x: x.day)
    sorted_daily_data = sorted(daily_data, key=lambda x: x.day, reverse=True)

    # Obtener datos de cierre del mes anterior (del día 1)
    first_day_data = next((d for d in sorted_daily_data if d.day == 1), None)

    for data in sorted_daily_data:
        # Obtener valores base del día actual
        mad = safe_decimal(data.mad)
        dia = safe_decimal(data.dia)
        pico = safe_decimal(data.pico)
        react = safe_decimal(data.react)

        # Para el día 1, usar valores de cierre del mes anterior como "día anterior"
        if data.day == 1:
            prev_mad = safe_decimal(first_day_data.mad_cierre if first_day_data else None)
            prev_dia = safe_decimal(first_day_data.dia_cierre if first_day_data else None)
            prev_pico = safe_decimal(first_day_data.pico_cierre if first_day_data else None)
            prev_react = safe_decimal(first_day_data.react_cierre if first_day_data else None)
            prev_total = prev_mad + prev_dia + prev_pico
        else:
            # Para otros días, obtener datos del día anterior
            prev_day = next((d for d in sorted_daily_data if d.day == data.day - 1), None)
            prev_mad = safe_decimal(prev_day.mad if prev_day else None)
            prev_dia = safe_decimal(prev_day.dia if prev_day else None)
            prev_pico = safe_decimal(prev_day.pico if prev_day else None)
            prev_react = safe_decimal(prev_day.react if prev_day else None)
            prev_total = prev_mad + prev_dia + prev_pico

        # 2.1) Diferencias (respecto al día anterior)
        diferencia_mad = mad - prev_mad if prev_mad is not None else mad
        diferencia_dia = dia - prev_dia if prev_dia is not None else dia
        diferencia_pico = pico - prev_pico if prev_pico is not None else pico
        diferencia_react = react - prev_react if prev_react is not None else react

        # 1.1) Cálculos básicos
        total_diario = mad + dia + pico

        # 4.1) Acumulados
        if data.day == 1:
            consumo_acumulado_diario = diferencia_mad + diferencia_dia + diferencia_pico + daily_loss
        else:
            consumo_acumulado_diario = diferencia_mad + diferencia_dia + diferencia_pico

        # consumo_con_pico = total_diario + ((mad / Decimal('24')) * Decimal(str(factor_pico)))
        # consumo_con_react = consumo_con_pico + ((dia / Decimal('12')) * Decimal(str(factor_react)))

        # 1.2) Cálculos básicos
        consumo_con_pico = (prev_total if prev_total is not None else Decimal('0')) + ((consumo_acumulado_diario / Decimal('24')) * Decimal(str(factor_pico)))
        consumo_con_react = consumo_con_pico + ((diferencia_dia / Decimal('12')) * Decimal(str(factor_react)))

        # 3.1) Cálculos de plan vs real
        pico_diario = diferencia_pico
        diferencia_plan_pico = daily_plan_pico - pico_diario
        diferencia_consumo = consumo_con_react - consumo_con_pico

        # 4.1) Acumulados
        diferencia_acumulada_diaria = daily_plan - consumo_acumulado_diario

        # Para acumulados totales, necesitamos los valores del día anterior
        if data.day == 1:
            # Usar valores de cierre del mes anterior si están disponibles
            if first_day_data:
                plan_acumulado_total = 0 + daily_plan_pico
                pico_acumulado_total = 0 + pico_diario
                plan_mensual = 0 + daily_plan
                consumo_mensual = 0 + consumo_acumulado_diario
            else:
                plan_acumulado_total = daily_plan_pico
                pico_acumulado_total = pico_diario
                plan_mensual = daily_plan
                consumo_mensual = consumo_acumulado_diario
        else:
            try:
                prev_day_data = TablaBitacoraData.objects.get(
                    area=area,
                    month=month,
                    year=year,
                    day=data.day - 1
                )
                plan_acumulado_total = prev_day_data.plan_acumulado_total + daily_plan_pico
                pico_acumulado_total = prev_day_data.pico_acumulado_total + pico_diario
                plan_mensual = prev_day_data.plan_mensual + daily_plan
                consumo_mensual = prev_day_data.consumo_mensual + consumo_acumulado_diario
            except TablaBitacoraData.DoesNotExist:
                plan_acumulado_total = daily_plan_pico
                pico_acumulado_total = pico_diario
                plan_mensual = daily_plan
                consumo_mensual = consumo_acumulado_diario

        diferencia_total = plan_acumulado_total - pico_acumulado_total
        diferencia_mensual = plan_mensual - consumo_mensual

        # 5. Factor de potencia
        try:
            factor_potencia = Decimal(math.cos(math.atan(float(diferencia_react) / float(consumo_acumulado_diario))))
        except:
            factor_potencia = Decimal('0')

        # Guardar en la base de datos
        TablaBitacoraData.objects.update_or_create(
            area=area,
            month=month,
            year=year,
            day=data.day,
            defaults={
                'mad': mad,
                'dia': dia,
                'pico': pico,
                'react': react,
                'total_diario': total_diario,
                'consumo_con_pico': consumo_con_pico,
                'consumo_con_react': consumo_con_react,
                'diferencia_mad': diferencia_mad,
                'diferencia_dia': diferencia_dia,
                'diferencia_pico': diferencia_pico,
                'diferencia_react': diferencia_react,
                'plan_diario': daily_plan_pico,
                'pico_diario': pico_diario,
                'diferencia_plan_pico': diferencia_plan_pico,
                'diferencia_consumo': diferencia_consumo,
                'plan_acumulado': daily_plan,
                'consumo_acumulado': consumo_acumulado_diario,
                'diferencia_acumulada': diferencia_acumulada_diaria,
                'plan_acumulado_total': plan_acumulado_total,
                'pico_acumulado_total': pico_acumulado_total,
                'diferencia_total': diferencia_total,
                'plan_mensual': plan_mensual,
                'consumo_mensual': consumo_mensual,
                'diferencia_mensual': diferencia_mensual,
                'factor_potencia': factor_potencia
            }
        )

    # Procesar cierre del mes anterior (fila 42)
    # cierre_row = 42
    first_day_data = next((d for d in daily_data if d.day == 1), None)
        
    # Obtener el último día del mes actual (sin modificaciones)
    last_day_data = next((d for d in daily_data if d.day == days_in_month), None)
    
    if last_day_data:
        # Obtener el consumo mensual del último registro de TablaBitacoraData
        try:
            last_day_bitacora = TablaBitacoraData.objects.get(
                area=area,
                month=month,
                year=year,
                day=days_in_month
            )
            consumo_mensual = last_day_bitacora.consumo_mensual
        except TablaBitacoraData.DoesNotExist:
            consumo_mensual = Decimal('0')

        # Calcular M (suma de diferencia_react) y AA (promedio de factor_potencia)
        sum_m = TablaBitacoraData.objects.filter(
            area=area,
            month=month,
            year=year
        ).aggregate(Sum('diferencia_react'))['diferencia_react__sum'] or Decimal('0')
        
        avg_aa = TablaBitacoraData.objects.filter(
            area=area,
            month=month,
            year=year
        ).aggregate(Avg('factor_potencia'))['factor_potencia__avg'] or Decimal('0')

        # Guardar cierre en la base de datos para el próximo mes
        TransformerLossData.objects.update_or_create(
            area=area,
            month=month,
            year=year,
            # month=month + 1 if month < 12 else 1,  # Mes siguiente
            # year=year if month < 12 else year + 1,  # Año siguiente si es diciembre
            defaults={
                'reactive_consumption': sum_m,
                'monthly_consumption': consumo_mensual,
            }
        )

        # Guardar datos de cierre para el próximo mes
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        
        # Crear o actualizar EnergyConsumption para el próximo mes (puede estar vacío)
        energy_consumption_next, _ = EnergyConsumption.objects.get_or_create(
            area=area,
            month=next_month,
            year=next_year,
    
            defaults={
                'plan_mes': 0,  # Puedes ajustar esto según necesidades
                'perdidas': 0    # Puedes ajustar esto según necesidades
            }
        )

        # Crear registro para el día 1 del próximo mes con datos de cierre
        DailyConsumption.objects.update_or_create(
            energy_consumption=energy_consumption_next,
            day=1,
            defaults={
                'mad_cierre': last_day_data.mad,
                'dia_cierre': last_day_data.dia,
                'pico_cierre': last_day_data.pico,
                'total_cierre': (Decimal(last_day_data.mad or 0) + 
                                Decimal(last_day_data.dia or 0) + 
                                Decimal(last_day_data.pico or 0)),
                'pico_diurno_11am_cierre': last_day_data.pico_diurno_11am_cierre,
                'pico_diurno_1pm_cierre': last_day_data.pico_diurno_1pm_cierre,
                'react_cierre': last_day_data.react
            }
        )

def fill_daily_data(ws, daily_data, month, year, factor_pico, factor_react):
    """Llena los datos diarios en la hoja de cálculo"""
    days = days_in_month(month, year)
    base_row = 41  # Fila base donde empezará el día 1
    
    for data in daily_data:
        row = base_row - data.day + 1  # Calcula la fila basada en el día
        
        # Escribir valores básicos
        ws[f'B{row}'] = data.day
        ws[f'C{row}'] = data.mad if data.mad else ''
        ws[f'D{row}'] = data.dia if data.dia else ''
        ws[f'E{row}'] = data.pico if data.pico else ''
        ws[f'I{row}'] = data.react if data.react else ''
        
        # Configurar fórmulas según el día
        if data.day == 1:
            # Fórmulas especiales para el día 1
            ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
            ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
            ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
            ws[f'J{row}'] = f'=C{row}-C{row+1}'
            ws[f'K{row}'] = f'=D{row}-D{row+1}'
            ws[f'L{row}'] = f'=E{row}-E{row+1}'
            ws[f'M{row}'] = f'=I{row}-I{row+1}'
            ws[f'N{row}'] = f'=O6/{days}'
            ws[f'O{row}'] = f'=L{row}'
            ws[f'P{row}'] = f'=N{row}-O{row}'
            ws[f'Q{row}'] = f'=H{row}-G{row}'
            ws[f'R{row}'] = f'=S6/{days}'
            ws[f'S{row}'] = f'=J{row}+K{row}+L{row}+V6'
            ws[f'T{row}'] = f'=R{row}-S{row}'
            ws[f'U{row}'] = f'=N{row}'
            ws[f'V{row}'] = f'=O{row}'
            ws[f'W{row}'] = f'=U{row}-V{row}'
            ws[f'X{row}'] = f'=R{row}'
            ws[f'Y{row}'] = f'=S{row}'
            ws[f'Z{row}'] = f'=X{row}-Y{row}'
            ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'
        else:
            # Fórmulas para días normales
            ws[f'F{row}'] = f'=C{row}+D{row}+E{row}'
            ws[f'G{row}'] = f'=(C{row+1}+D{row+1}+E{row+1})+((S{row}/24)*{factor_pico})'
            ws[f'H{row}'] = f'=G{row}+((K{row}/12)*{factor_react})'
            ws[f'J{row}'] = f'=C{row}-C{row+1}'
            ws[f'K{row}'] = f'=D{row}-D{row+1}'
            ws[f'L{row}'] = f'=E{row}-E{row+1}'
            ws[f'M{row}'] = f'=I{row}-I{row+1}'
            ws[f'N{row}'] = f'=N{row+1}'
            ws[f'O{row}'] = f'=L{row}'
            ws[f'P{row}'] = f'=N{row}-O{row}'
            ws[f'Q{row}'] = f'=H{row}-G{row}'
            ws[f'R{row}'] = f'=R{row+1}'
            ws[f'S{row}'] = f'=J{row}+K{row}+L{row}'
            ws[f'T{row}'] = f'=R{row}-S{row}'
            ws[f'U{row}'] = f'=U{row+1}+N{row}'
            ws[f'V{row}'] = f'=V{row+1}+O{row}'
            ws[f'W{row}'] = f'=U{row}-V{row}'
            ws[f'X{row}'] = f'=X{row+1}+R{row}'
            ws[f'Y{row}'] = f'=Y{row+1}+S{row}'
            ws[f'Z{row}'] = f'=X{row}-Y{row}'
            ws[f'AA{row}'] = f'=COS(ATAN(M{row}/S{row}))'

        # 2. Luego calculamos y guardamos los valores directos en la base de datos
        calculate_and_save_direct_values(ws, daily_data, month, year, factor_pico, factor_react)
    
        # Resaltar celdas importantes
        for col in ['N41', 'R41']:
            ws[f'{col}{row}'].fill = YELLOW_FILL
    
    # Fila de cierre del mes anterior
    cierre_row = 42
    if daily_data.exists():
        last_day_data = daily_data.last()
        ws[f'B{cierre_row}'] = 'cierre mes anterior'
        ws[f'C{cierre_row}'] = last_day_data.mad_cierre if last_day_data.mad_cierre else ''
        ws[f'D{cierre_row}'] = last_day_data.dia_cierre if last_day_data.dia_cierre else ''
        ws[f'E{cierre_row}'] = last_day_data.pico_cierre if last_day_data.pico_cierre else ''
        ws[f'F{cierre_row}'] = last_day_data.total_cierre if last_day_data.total_cierre else f'=C{cierre_row}+D{cierre_row}+E{cierre_row}'
        ws[f'G{cierre_row}'] = last_day_data.pico_diurno_11am_cierre if last_day_data.pico_diurno_11am_cierre else ''
        ws[f'H{cierre_row}'] = last_day_data.pico_diurno_1pm_cierre if last_day_data.pico_diurno_1pm_cierre else ''
        ws[f'I{cierre_row}'] = last_day_data.react_cierre if last_day_data.react_cierre else ''
    else:
        ws[f'B{cierre_row}'] = 'cierre mes anterior'
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{cierre_row}'] = ''

    # Calcular el rango de filas con datos
    first_data_row = base_row - days + 1
    ws[f'M{cierre_row}'] = f'=SUM(M{first_data_row}:M{base_row})'
    ws[f'AA{cierre_row}'] = f'=AVERAGE(AA{first_data_row}:AA{base_row})'

def days_in_month(month, year):
    from calendar import monthrange
    return monthrange(year, month)[1]

